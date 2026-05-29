from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from src.eval.composition_qos_eval import evaluate_composition_qos


class CompositionQosEvalTests(unittest.TestCase):
    def _write_json(self, path: Path, payload) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(payload), encoding="utf-8")

    def test_average_availability_and_completeness_score_gate(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            query_dir = Path(tmp)
            eval_dir = query_dir / "evaluation"
            self._write_json(
                query_dir / "0_decomposer.json",
                [
                    {"id": 1, "description": "First API call"},
                    {"id": 2, "description": "Second API call"},
                ],
            )
            selected_rows = {
                "api_a": {"api_id": "api_a", "service": {"qos": {"rt_s": 10, "tp_kbps": 50, "availability": 0.9}}},
                "api_b": {"api_id": "api_b", "service": {"qos": {"rt_s": 20, "tp_kbps": 40, "availability": 0.8}}},
            }
            self._write_json(query_dir / "no_qos" / "3_selected_s1.json", [selected_rows["api_a"]])
            self._write_json(query_dir / "no_qos" / "3_selected_s2.json", [selected_rows["api_b"]])
            self._write_json(query_dir / "qos_pure_llm" / "3_selected_s1.json", [selected_rows["api_a"]])
            self._write_json(
                query_dir / "no_qos" / "4_planner.json",
                {
                    "primary_plan": {
                        "plan_id": 1,
                        "summary": "Valid plan",
                        "steps": [
                            {
                                "step": 1,
                                "api_id": "api_a",
                                "subtask_id": 1,
                                "action": "Call A",
                                "input_from_previous_step": None,
                                "output_to_next_step": "a result",
                                "why": "Matches",
                                "qos": None,
                            },
                            {
                                "step": 2,
                                "api_id": "api_b",
                                "subtask_id": 2,
                                "action": "Call B",
                                "input_from_previous_step": "a result",
                                "output_to_next_step": "b result",
                                "why": "Matches",
                                "qos": None,
                            },
                        ],
                        "subtask_coverage": [],
                    },
                    "selected_api_ids": ["api_a", "api_b"],
                    "overall_rationale": "Valid",
                },
            )
            self._write_json(
                query_dir / "qos_pure_llm" / "4_planner.json",
                {
                    "primary_plan": {
                        "plan_id": 1,
                        "summary": "Incomplete but structurally valid plan",
                        "steps": [
                            {
                                "step": 1,
                                "api_id": "api_a",
                                "subtask_id": 1,
                                "action": "Call A",
                                "input_from_previous_step": None,
                                "output_to_next_step": "a result",
                                "why": "Matches",
                                "qos": None,
                            }
                        ],
                        "subtask_coverage": [],
                    },
                    "selected_api_ids": ["api_a"],
                    "overall_rationale": "Incomplete",
                },
            )
            self._write_json(
                eval_dir / "query_qx_candidate_api_rankings_rows.json",
                [
                    {
                        "Mode": "no_qos",
                        "Sub Task": "1",
                        "Selected_API": "api_a",
                        "Functional Match (0/1)": 1,
                        "QoS_RT_s": 10,
                        "QoS_TP_kbps": 50,
                        "QoS Availability": 0.9,
                    },
                    {
                        "Mode": "no_qos",
                        "Sub Task": "2",
                        "Selected_API": "api_b",
                        "Functional Match (0/1)": 1,
                        "QoS_RT_s": 20,
                        "QoS_TP_kbps": 40,
                        "QoS Availability": 0.8,
                    },
                    {
                        "Mode": "qos_pure_llm",
                        "Sub Task": "1",
                        "Selected_API": "api_a",
                        "Functional Match (0/1)": 1,
                        "QoS_RT_s": 40,
                        "QoS_TP_kbps": 10,
                        "QoS Availability": 0.5,
                    },
                ],
            )

            result = evaluate_composition_qos(query_dir=query_dir, query_id="qx", output_dir=eval_dir)
            rows_by_mode = {row["Mode"]: row for row in result["rows"]}

            self.assertEqual(rows_by_mode["no_qos"]["Composition_Validity"], 1)
            self.assertEqual(rows_by_mode["no_qos"]["Composition_Completeness_Gate"], 1.0)
            self.assertEqual(rows_by_mode["no_qos"]["Average_Workflow_Availability"], 0.85)
            self.assertEqual(rows_by_mode["no_qos"]["QoS_Adjusted_Composition_Score"], 1.0)
            self.assertEqual(rows_by_mode["qos_pure_llm"]["Composition_Validity"], 1)
            self.assertEqual(rows_by_mode["qos_pure_llm"]["Composition_Completeness"], 0.0)
            self.assertEqual(rows_by_mode["qos_pure_llm"]["Composition_Completeness_Gate"], 0.0)
            self.assertEqual(rows_by_mode["qos_pure_llm"]["QoS_Adjusted_Composition_Score"], 0.0)
            self.assertEqual(rows_by_mode["qos_topsis"]["Composition_Validity"], 0)
            self.assertEqual(rows_by_mode["qos_topsis"]["QoS_Adjusted_Composition_Score"], 0.0)
            self.assertTrue(Path(result["rows_json"]).exists())
            self.assertTrue(Path(result["summary_json"]).exists())
            self.assertTrue(Path(result["excel"]).exists())

            summary = json.loads(Path(result["summary_json"]).read_text(encoding="utf-8"))
            self.assertEqual(summary["complete_mode_count"], 1)
            self.assertIn("Composition_Validity is diagnostic only", summary["scoring_formula"])
            workbook = load_workbook(result["excel"], read_only=True)
            definitions = {
                row[0]: row[2]
                for row in workbook["Metric_Definitions"].iter_rows(min_row=2, values_only=True)
            }
            self.assertIn("Average_Workflow_Availability", definitions)
            self.assertIn("0.7 * Functional_Coverage + 0.3 * Normalized_QoS_Score", definitions["QoS_Adjusted_Composition_Score"])

    def test_summary_reports_all_tied_best_modes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            query_dir = Path(tmp)
            eval_dir = query_dir / "evaluation"
            self._write_json(query_dir / "0_decomposer.json", [{"id": 1, "description": "Call API"}])
            selected = {"api_id": "api_a", "service": {"qos": {"rt_s": 10, "tp_kbps": 50, "availability": 0.9}}}
            for mode in ["no_qos", "qos_hybrid"]:
                self._write_json(query_dir / mode / "3_selected_s1.json", [selected])
                self._write_json(
                    query_dir / mode / "4_planner.json",
                    {
                        "primary_plan": {
                            "plan_id": 1,
                            "summary": "Valid plan",
                            "steps": [
                                {
                                    "step": 1,
                                    "api_id": "api_a",
                                    "subtask_id": 1,
                                    "action": "Call A",
                                    "input_from_previous_step": None,
                                    "output_to_next_step": "a result",
                                    "why": "Matches",
                                    "qos": None,
                                }
                            ],
                            "subtask_coverage": [],
                        },
                        "selected_api_ids": ["api_a"],
                        "overall_rationale": "Valid",
                    },
                )
            self._write_json(
                eval_dir / "query_qtie_candidate_api_rankings_rows.json",
                [
                    {
                        "Mode": mode,
                        "Sub Task": "1",
                        "Selected_API": "api_a",
                        "Functional Match (0/1)": 1,
                        "QoS_RT_s": 10,
                        "QoS_TP_kbps": 50,
                        "QoS Availability": 0.9,
                    }
                    for mode in ["no_qos", "qos_hybrid"]
                ],
            )

            result = evaluate_composition_qos(query_dir=query_dir, query_id="qtie", output_dir=eval_dir)
            summary = json.loads(Path(result["summary_json"]).read_text(encoding="utf-8"))

            self.assertEqual(summary["best_modes_by_qos_adjusted_composition_score"], ["no_qos", "qos_hybrid"])
            self.assertEqual(summary["best_qos_adjusted_composition_score"], 1.0)
            self.assertTrue(summary["is_qos_adjusted_composition_score_tie"])
            self.assertEqual(summary["best_mode_by_qos_adjusted_composition_score"]["mode"], "no_qos")
            summary_rows = {row["Mode"]: row for row in summary["summary_rows"]}
            self.assertEqual(summary_rows["no_qos"]["Short_Interpretation"], "Tied best overall composition score")
            self.assertEqual(summary_rows["qos_hybrid"]["Short_Interpretation"], "Tied best overall composition score")

            workbook = load_workbook(result["excel"], read_only=True)
            best_rows = {
                row[0]: row
                for row in workbook["Best_Mode_Summary"].iter_rows(min_row=2, values_only=True)
            }
            qos_row = best_rows["QoS_Adjusted_Composition_Score"]
            self.assertEqual(qos_row[3], "no_qos, qos_hybrid")
            self.assertEqual(qos_row[4], "Yes")

    def test_missing_qos_still_scores_functional_component(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            query_dir = Path(tmp)
            eval_dir = query_dir / "evaluation"
            self._write_json(query_dir / "0_decomposer.json", [{"id": 1, "description": "Call API"}])
            self._write_json(
                query_dir / "no_qos" / "3_selected_s1.json",
                [{"api_id": "api_missing_qos", "service": {"qos": {"rt_s": None, "tp_kbps": None, "availability": 0.0}}}],
            )
            self._write_json(
                query_dir / "no_qos" / "4_planner.json",
                {
                    "primary_plan": {
                        "plan_id": 1,
                        "summary": "Complete plan with missing QoS",
                        "steps": [
                            {
                                "step": 1,
                                "api_id": "api_missing_qos",
                                "subtask_id": 1,
                                "action": "Call API",
                                "input_from_previous_step": None,
                                "output_to_next_step": "result",
                                "why": "Matches",
                                "qos": None,
                            }
                        ],
                        "subtask_coverage": [],
                    },
                    "selected_api_ids": ["api_missing_qos"],
                    "overall_rationale": "Valid",
                },
            )
            self._write_json(
                eval_dir / "query_qmissing_candidate_api_rankings_rows.json",
                [
                    {
                        "Mode": "no_qos",
                        "Sub Task": "1",
                        "Selected_API": "api_missing_qos",
                        "Functional Match (0/1)": 1,
                        "QoS_RT_s": None,
                        "QoS_TP_kbps": None,
                        "QoS Availability": 0,
                    }
                ],
            )

            result = evaluate_composition_qos(query_dir=query_dir, query_id="qmissing", output_dir=eval_dir)
            rows_by_mode = {row["Mode"]: row for row in result["rows"]}

            self.assertEqual(rows_by_mode["no_qos"]["Composition_Completeness"], 1.0)
            self.assertEqual(rows_by_mode["no_qos"]["Functional_Coverage"], 1.0)
            self.assertEqual(rows_by_mode["no_qos"]["Normalized_QoS_Score"], 0.0)
            self.assertEqual(rows_by_mode["no_qos"]["QoS_Adjusted_Composition_Score"], 0.7)

    def test_qos_references_use_max_functional_coverage_complete_workflows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            query_dir = Path(tmp)
            eval_dir = query_dir / "evaluation"
            self._write_json(
                query_dir / "0_decomposer.json",
                [
                    {"id": 1, "description": "First API call"},
                    {"id": 2, "description": "Second API call"},
                ],
            )

            selected_by_mode = {
                "no_qos": [
                    ("api_no_1", 10, 100, 0.9),
                    ("api_no_2", 10, 100, 0.9),
                ],
                "qos_hybrid": [
                    ("api_hybrid_1", 10, 200, 0.9),
                    ("api_hybrid_2", 10, 200, 0.9),
                ],
                "qos_topsis": [
                    ("api_topsis_1", 10, 1000, 0.9),
                    ("api_topsis_2", 10, 1000, 0.9),
                ],
            }

            candidate_rows = []
            for mode, api_specs in selected_by_mode.items():
                steps = []
                selected_api_ids = []
                for subtask_id, (api_id, rt_s, tp_kbps, availability) in enumerate(api_specs, start=1):
                    selected_api_ids.append(api_id)
                    self._write_json(
                        query_dir / mode / f"3_selected_s{subtask_id}.json",
                        [
                            {
                                "api_id": api_id,
                                "service": {"qos": {"rt_s": rt_s, "tp_kbps": tp_kbps, "availability": availability}},
                            }
                        ],
                    )
                    steps.append(
                        {
                            "step": subtask_id,
                            "api_id": api_id,
                            "subtask_id": subtask_id,
                            "action": f"Call {api_id}",
                            "input_from_previous_step": None if subtask_id == 1 else "previous result",
                            "output_to_next_step": "result",
                            "why": "Matches",
                            "qos": None,
                        }
                    )
                    candidate_rows.append(
                        {
                            "Mode": mode,
                            "Sub Task": str(subtask_id),
                            "Selected_API": api_id,
                            "Functional Match (0/1)": 0 if mode == "qos_topsis" else 1,
                            "QoS_RT_s": rt_s,
                            "QoS_TP_kbps": tp_kbps,
                            "QoS Availability": availability,
                        }
                    )

                self._write_json(
                    query_dir / mode / "4_planner.json",
                    {
                        "primary_plan": {
                            "plan_id": 1,
                            "summary": "Complete plan",
                            "steps": steps,
                            "subtask_coverage": [],
                        },
                        "selected_api_ids": selected_api_ids,
                        "overall_rationale": "Complete",
                    },
                )

            self._write_json(eval_dir / "query_qref_candidate_api_rankings_rows.json", candidate_rows)

            result = evaluate_composition_qos(query_dir=query_dir, query_id="qref", output_dir=eval_dir)
            rows_by_mode = {row["Mode"]: row for row in result["rows"]}

            self.assertEqual(rows_by_mode["qos_topsis"]["Composition_Completeness"], 1.0)
            self.assertEqual(rows_by_mode["qos_topsis"]["Functional_Coverage"], 0.0)
            self.assertEqual(rows_by_mode["qos_topsis"]["Normalized_QoS_Score"], 1.0)
            self.assertEqual(rows_by_mode["qos_topsis"]["QoS_Adjusted_Composition_Score"], 0.3)
            self.assertEqual(rows_by_mode["qos_hybrid"]["Normalized_Throughput_Score"], 1.0)
            self.assertEqual(rows_by_mode["no_qos"]["Normalized_Throughput_Score"], 0.5)


if __name__ == "__main__":
    unittest.main()
