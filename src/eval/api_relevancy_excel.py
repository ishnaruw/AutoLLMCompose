from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

COLUMNS = [
    "Query_ID",
    "Mode",
    "Sub Task",
    "Retrieved Rank",
    "Mode Rank",
    "Subtask_Purpose",
    "Selected_API",
    "Is Hallucinated? (0/1)",
    "Is Duplicated? (0/1)",
    "Functional Match Label",
    "Used in Ranking",
    "Selected for Planner",
    "Planner Selection K",
    "QoS_RT",
    "QoS_TP",
    "QoS Availability",
    "Comments",
]

MODE_ORDER = ["no_qos", "qos_pure_llm", "qos_topsis", "qos_hybrid"]
MODE_INDEX = {name: idx for idx, name in enumerate(MODE_ORDER)}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SUMMARY_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")


def _subtask_sort_key(value: Any) -> Tuple[int, str]:
    text = str(value)
    return (int(text), text) if text.isdigit() else (9999, text)


def _mode_sort_key(value: Any) -> Tuple[int, str]:
    text = str(value)
    return MODE_INDEX.get(text, 9999), text


def _rank_sort_key(value: Any) -> int:
    try:
        return int(value)
    except Exception:
        return 9999


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except Exception:
        return default


def _functional_match_label(row: Dict[str, Any]) -> int:
    return _safe_int(
        row.get(
            "Functional Match Label",
            row.get("API Relevancy (0/1)", row.get("relevant", 0)),
        )
    )


def _normalize_flag(value: Any) -> int | None:
    if value in (0, 1):
        return int(value)
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, str):
        text = value.strip().lower()
        if text in {"0", "1"}:
            return int(text)
        if text in {"true", "yes", "y"}:
            return 1
        if text in {"false", "no", "n"}:
            return 0
    return None


def _split_csv_like(value: Any) -> List[str]:
    if value is None:
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def _build_duplicate_flag_keys(duplicate_audit: Dict[str, Any] | None) -> set[Tuple[str, str, str, str]]:
    if not duplicate_audit:
        return set()

    flagged: set[Tuple[str, str, str, str]] = set()

    for row in duplicate_audit.get("duplicate_rows", []):
        flagged.add(
            (
                str(row.get("Query_ID", "")),
                str(row.get("Sub_Task", "")),
                str(row.get("Mode", "")),
                str(row.get("Selected_API", "")),
            )
        )

    for row in duplicate_audit.get("summary_rows", []):
        query_id = str(row.get("Query_ID", ""))
        subtask_id = str(row.get("Sub_Task", ""))
        mode = str(row.get("Mode", ""))
        for part in _split_csv_like(row.get("Duplicated_APIs", "")):
            api_id = part.split(" x", 1)[0].strip()
            if api_id:
                flagged.add((query_id, subtask_id, mode, api_id))

    return flagged


def _build_hallucination_flag_keys(hallucination_audit: Dict[str, Any] | None) -> set[Tuple[str, str, str, str]]:
    if not hallucination_audit:
        return set()

    flagged: set[Tuple[str, str, str, str]] = set()

    for row in hallucination_audit.get("mode_detail_rows", []):
        flagged.add(
            (
                str(row.get("Query_ID", "")),
                str(row.get("Sub_Task", "")),
                str(row.get("Mode", "")),
                str(row.get("Selected_API", "")),
            )
        )

    for row in hallucination_audit.get("mode_summary_rows", []):
        query_id = str(row.get("Query_ID", ""))
        subtask_id = str(row.get("Sub_Task", ""))
        mode = str(row.get("Mode", ""))
        for api_id in _split_csv_like(row.get("Hallucinated_APIs", "")):
            flagged.add((query_id, subtask_id, mode, api_id))

    return flagged


def enrich_relevancy_rows_with_anomaly_flags(
    rows: List[Dict[str, Any]],
    *,
    duplicate_audit: Dict[str, Any] | None = None,
    hallucination_audit: Dict[str, Any] | None = None,
) -> List[Dict[str, Any]]:
    duplicate_keys = _build_duplicate_flag_keys(duplicate_audit)
    hallucination_keys = _build_hallucination_flag_keys(hallucination_audit)

    enriched_rows: List[Dict[str, Any]] = []
    for row in rows:
        enriched = dict(row)
        if "Functional Match Label" not in enriched and "API Relevancy (0/1)" in enriched:
            enriched["Functional Match Label"] = enriched.get("API Relevancy (0/1)")
        key = (
            str(row.get("Query_ID", "")),
            str(row.get("Sub Task", "")),
            str(row.get("Mode", "")),
            str(row.get("Selected_API", "")),
        )

        duplicate_flag = _normalize_flag(row.get("Is Duplicated? (0/1)"))
        hallucination_flag = _normalize_flag(row.get("Is Hallucinated? (0/1)"))

        if duplicate_flag is None and duplicate_audit is not None:
            duplicate_flag = 1 if key in duplicate_keys else 0
        if hallucination_flag is None and hallucination_audit is not None:
            hallucination_flag = 1 if key in hallucination_keys else 0

        if duplicate_flag is not None:
            enriched["Is Duplicated? (0/1)"] = duplicate_flag
        if hallucination_flag is not None:
            enriched["Is Hallucinated? (0/1)"] = hallucination_flag

        enriched_rows.append(enriched)

    return enriched_rows


def _style_header(ws, columns: List[str]) -> None:
    ws.append(columns)
    for idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _finalize_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 60)


def _append_sheet(
    ws,
    columns: List[str],
    rows: List[Dict[str, Any]],
    *,
    highlight_fn=None,
) -> None:
    _style_header(ws, columns)

    for row in rows:
        ws.append([row.get(column, "") for column in columns])
        fill = highlight_fn(row) if highlight_fn else None
        if fill is not None:
            for idx in range(1, len(columns) + 1):
                ws.cell(row=ws.max_row, column=idx).fill = fill

    _finalize_sheet(ws)


def _append_legacy_query_sheet(
    ws,
    rows: List[Dict[str, Any]],
) -> None:
    _style_header(ws, COLUMNS)

    grouped: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[(str(row.get("Sub Task", "")), str(row.get("Mode", "")))].append(row)

    subtasks = sorted({str(row.get("Sub Task", "")) for row in rows}, key=_subtask_sort_key)
    for subtask_id in subtasks:
        for mode in MODE_ORDER:
            group = sorted(
                grouped.get((subtask_id, mode), []),
                key=lambda row: _rank_sort_key(row.get("Mode Rank")),
            )
            if not group:
                continue

            for row in group:
                ws.append([row.get(column, "") for column in COLUMNS])

            relevant = sum(1 for row in group if _functional_match_label(row) == 1)
            total = len(group)
            precision = round(relevant / float(total or 1), 4)

            summary_row = {
                "Query_ID": group[0].get("Query_ID"),
                "Mode": mode,
                "Sub Task": subtask_id,
                "Subtask_Purpose": group[0].get("Subtask_Purpose", ""),
                "Selected_API": "Precision",
                "Functional Match Label": precision,
                "Comments": f"Precision = {relevant}/{total}",
            }
            ws.append([summary_row.get(column, "") for column in COLUMNS])
            for idx in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=ws.max_row, column=idx)
                cell.font = Font(bold=True)
                cell.fill = SUMMARY_FILL

            ws.append([""] * len(COLUMNS))

    _finalize_sheet(ws)


def _build_precision_rows(
    rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[(str(row.get("Query_ID", "")), str(row.get("Sub Task", "")), str(row.get("Mode", "")))].append(row)

    precision_rows: List[Dict[str, Any]] = []
    for (query_id, subtask_id, mode), group in sorted(
        grouped.items(),
        key=lambda item: (_subtask_sort_key(item[0][1]), _mode_sort_key(item[0][2])),
    ):
        relevant = sum(1 for row in group if _functional_match_label(row) == 1)
        total = len(group)
        precision_rows.append(
            {
                "Query_ID": query_id,
                "Sub Task": subtask_id,
                "Mode": mode,
                "Subtask_Purpose": str(group[0].get("Subtask_Purpose", "")) if group else "",
                "Functional_Match_Count": relevant,
                "Candidate_Count": total,
                "Precision": round(relevant / float(total or 1), 4),
            }
        )
    return precision_rows


def _aggregate_mode_rows(precision_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in precision_rows:
        grouped[str(row.get("Mode", ""))].append(row)

    out: List[Dict[str, Any]] = []
    for mode, rows in sorted(grouped.items(), key=lambda item: _mode_sort_key(item[0])):
        relevant = sum(_safe_int(row.get("Functional_Match_Count", row.get("Relevant_Count"))) for row in rows)
        total = sum(_safe_int(row.get("Candidate_Count")) for row in rows)
        out.append(
            {
                "Mode": mode,
                "Subtasks": len(rows),
                "Functional_Match_Count": relevant,
                "Candidate_Count": total,
                "Precision": round(relevant / float(total or 1), 4),
            }
        )
    return out


def _aggregate_subtask_rows(precision_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in precision_rows:
        grouped[str(row.get("Sub Task", ""))].append(row)

    out: List[Dict[str, Any]] = []
    for subtask_id, rows in sorted(grouped.items(), key=lambda item: _subtask_sort_key(item[0])):
        relevant = sum(_safe_int(row.get("Functional_Match_Count", row.get("Relevant_Count"))) for row in rows)
        total = sum(_safe_int(row.get("Candidate_Count")) for row in rows)
        out.append(
            {
                "Sub Task": subtask_id,
                "Subtask_Purpose": str(rows[0].get("Subtask_Purpose", "")) if rows else "",
                "Modes": len(rows),
                "Functional_Match_Count": relevant,
                "Candidate_Count": total,
                "Precision": round(relevant / float(total or 1), 4),
            }
        )
    return out


def write_relevancy_excel(
    rows: List[Dict[str, Any]],
    out_path: str | Path,
    *,
    duplicate_audit: Dict[str, Any] | None = None,
    hallucination_audit: Dict[str, Any] | None = None,
) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    rows = enrich_relevancy_rows_with_anomaly_flags(
        rows,
        duplicate_audit=duplicate_audit,
        hallucination_audit=hallucination_audit,
    )

    ranked_rows = sorted(
        rows,
        key=lambda r: (
            _subtask_sort_key(r.get("Sub Task")),
            _mode_sort_key(r.get("Mode")),
            _rank_sort_key(r.get("Mode Rank")),
        ),
    )
    precision_rows = _build_precision_rows(rows)
    mode_summary_rows = _aggregate_mode_rows(precision_rows)
    subtask_summary_rows = _aggregate_subtask_rows(precision_rows)

    total_relevant = sum(_safe_int(row.get("Functional_Match_Count", row.get("Relevant_Count"))) for row in precision_rows)
    total_candidates = sum(_safe_int(row.get("Candidate_Count")) for row in precision_rows)

    wb = Workbook()

    query_ws = wb.active
    query_ws.title = "Query"
    _append_legacy_query_sheet(query_ws, ranked_rows)

    overview_ws = wb.create_sheet("Overview")
    overview_ws.title = "Overview"
    overview_rows = [
        {"Metric": "Query_ID", "Value": str(rows[0].get("Query_ID", "")) if rows else ""},
        {"Metric": "Total Ranked Rows", "Value": len(rows)},
        {"Metric": "Functional Match Rows", "Value": total_relevant},
        {"Metric": "Overall Precision", "Value": round(total_relevant / float(total_candidates or 1), 4)},
        {"Metric": "Subtasks", "Value": len({str(row.get("Sub Task", "")) for row in rows})},
        {"Metric": "Modes", "Value": len({str(row.get("Mode", "")) for row in rows})},
    ]
    _append_sheet(overview_ws, ["Metric", "Value"], overview_rows)

    mode_ws = wb.create_sheet("Mode Summary")
    _append_sheet(
        mode_ws,
        [
            "Mode",
            "Subtasks",
            "Functional_Match_Count",
            "Candidate_Count",
            "Precision",
        ],
        mode_summary_rows,
    )

    subtask_ws = wb.create_sheet("Subtask Summary")
    _append_sheet(
        subtask_ws,
        [
            "Sub Task",
            "Subtask_Purpose",
            "Modes",
            "Functional_Match_Count",
            "Candidate_Count",
            "Precision",
        ],
        subtask_summary_rows,
    )

    precision_ws = wb.create_sheet("Precision")
    _append_sheet(
        precision_ws,
        [
            "Query_ID",
            "Sub Task",
            "Mode",
            "Subtask_Purpose",
            "Functional_Match_Count",
            "Candidate_Count",
            "Precision",
        ],
        precision_rows,
    )

    ranked_ws = wb.create_sheet("Ranked APIs")
    _append_sheet(
        ranked_ws,
        COLUMNS,
        ranked_rows,
    )

    wb.active = 0
    wb.views[0].activeTab = 0
    wb.save(out_path)
    return out_path
