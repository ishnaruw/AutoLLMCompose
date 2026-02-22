"""src/eval/export_excel.py

Generate a per-run Excel report summarizing MAOF pipeline outputs.

Output:
  - One workbook per run folder (results/logs/<model>/<run_id>/report.xlsx)
  - One worksheet per query/run (single sheet named "Query")

The sheet contains four sections:
  1) Summary
  2) Subtasks
  3) Plans
  4) Selected (Top 8 per subtask)
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font


def _safe_read_json(path: Path) -> Any:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _bold_row(ws, row_idx: int) -> None:
    for cell in ws[row_idx]:
        cell.font = Font(bold=True)


def _append_blank(ws, n: int = 1) -> None:
    for _ in range(n):
        ws.append([""])


def export_run_excel(
    *,
    run_dir: str | Path,
    out_xlsx: Optional[str | Path] = None,
) -> Path:
    """Create an Excel report from a run folder."""

    run_dir = Path(run_dir)
    meta = _safe_read_json(run_dir / "meta.json") or {}
    subtasks = _safe_read_json(run_dir / "0_decomposer.json") or []
    retrieved_by_subtask = _safe_read_json(run_dir / "1_retriever.json") or {}
    ranked_pool = _safe_read_json(run_dir / "3_ranked_pool_with_service.json") or []
    selected = (
        _safe_read_json(run_dir / "4_selected_with_service_pure_llm.json")
        or _safe_read_json(run_dir / "4_selected_with_service.json")
        or []
    )
    plan = (
        _safe_read_json(run_dir / "5_planner_pure_llm.json")
        or _safe_read_json(run_dir / "5_planner.json")
        or {}
    )
    selector_traces = _safe_read_json(run_dir / "selector_traces.json") or {}

    if out_xlsx is None:
        out_xlsx = run_dir / "report.xlsx"
    out_xlsx = Path(out_xlsx)

    wb = Workbook()
    ws = wb.active
    ws.title = "Query"

    # ---------------- Summary ----------------
    ws.append(["Summary"])
    _bold_row(ws, ws.max_row)
    ws.append(["Run directory", str(run_dir)])
    ws.append(["Model tag", meta.get("model_tag")])
    ws.append(["Provider", meta.get("provider")])
    ws.append(["Model", meta.get("model")])
    ws.append(["with_qos", meta.get("with_qos")])
    ws.append(["Selector mode", meta.get("selector_mode")])
    ws.append(["Ranker pool N", meta.get("ranker_pool_n")])
    ws.append(["Selected top N", meta.get("selected_top_n")])
    ws.append(["RAG index dir", meta.get("rag_index_dir")])
    ws.append(["RAG topK", meta.get("rag_topk")])
    ws.append(["User goal", meta.get("user_goal")])

    _append_blank(ws, 2)

    # ---------------- Subtasks ----------------
    ws.append(["Subtasks"])
    _bold_row(ws, ws.max_row)
    ws.append([
        "subtask_id",
        "subtask",
        "retrieved_k",
        "ranker_pool_k",
        "selected_n",
        "selector_mode",
        "topsis_status",
    ])
    _bold_row(ws, ws.max_row)

    subtasks_list = subtasks if isinstance(subtasks, list) else []
    for st in subtasks_list:
        sid = str(st.get("id"))
        desc = st.get("description")
        retrieved = retrieved_by_subtask.get(sid, []) if isinstance(retrieved_by_subtask, dict) else []
        retrieved_k = len(retrieved) if isinstance(retrieved, list) else 0

        pool_k = len([c for c in ranked_pool if str(c.get("subtask_id")) == sid]) if isinstance(ranked_pool, list) else 0
        sel_k = len([c for c in selected if str(c.get("subtask_id")) == sid]) if isinstance(selected, list) else 0

        tr = selector_traces.get(sid, {}) if isinstance(selector_traces, dict) else {}
        ws.append([
            sid,
            desc,
            retrieved_k,
            pool_k,
            sel_k,
            tr.get("mode") or meta.get("selector_mode"),
            tr.get("topsis_status", "") if tr else "",
        ])

    _append_blank(ws, 2)

    # ---------------- Plans ----------------
    ws.append(["Plans"])
    _bold_row(ws, ws.max_row)
    ws.append([
        "path_id",
        "path_score",
        "step",
        "subtask_id",
        "api_id",
        "action",
        "why",
    ])
    _bold_row(ws, ws.max_row)

    paths = (plan.get("paths") or []) if isinstance(plan, dict) else []
    if isinstance(paths, list):
        for p in paths:
            pid = p.get("path_id")
            pscore = p.get("path_score")
            steps = p.get("steps") or []
            if not isinstance(steps, list):
                continue
            for s in steps:
                ws.append([
                    pid,
                    pscore,
                    s.get("step"),
                    s.get("subtask_id"),
                    s.get("api_id"),
                    s.get("action"),
                    s.get("why"),
                ])

    _append_blank(ws, 2)

    # ---------------- Selected APIs ----------------
    ws.append(["Selected APIs (Top N per subtask)"])
    _bold_row(ws, ws.max_row)
    ws.append([
        "subtask_id",
        "selected_rank",
        "api_id",
        "rag_score",
        "topsis_score",
        "rt_ms",
        "tp_rps",
        "availability",
    ])
    _bold_row(ws, ws.max_row)

    sel_list = selected if isinstance(selected, list) else []
    sel_list = sorted(sel_list, key=lambda x: (str(x.get("subtask_id")), int(x.get("selected_rank", 10**9))))
    for c in sel_list:
        svc = c.get("service") or {}
        qos = (svc.get("qos") or {}) if isinstance(svc, dict) else {}
        ws.append([
            c.get("subtask_id"),
            c.get("selected_rank"),
            c.get("api_id"),
            c.get("rag_score"),
            c.get("topsis_score", ""),
            qos.get("rt_ms"),
            qos.get("tp_rps"),
            qos.get("availability"),
        ])

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    return out_xlsx
