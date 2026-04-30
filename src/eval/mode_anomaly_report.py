from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

COLUMNS = [
    "Query_ID",
    "Subtask",
    "Mode",
    "APIs_Hallucinated",
    "APIs_Duplicated",
    "Ranking_Anomaly",
    "Ranking_Anomaly_Reason",
    "Missing_API_IDs",
    "Unknown_API_IDs",
    "Duplicate_API_IDs",
    "Expected_API_Count",
    "Actual_API_Count",
    "Returned_API_Count",
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
ALERT_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
MODE_ORDER = ["no_qos", "qos_pure_llm", "qos_topsis", "qos_hybrid"]
MODE_INDEX = {name: idx for idx, name in enumerate(MODE_ORDER)}


def _query_sort_key(value: Any) -> Tuple[int, str]:
    text = str(value)
    if text.startswith("q") and text[1:].isdigit():
        return int(text[1:]), text
    return 9999, text


def _subtask_sort_key(value: Any) -> Tuple[int, str]:
    text = str(value)
    return (int(text), text) if text.isdigit() else (9999, text)


def _mode_sort_key(value: Any) -> Tuple[int, str]:
    text = str(value)
    return MODE_INDEX.get(text, 9999), text


def _split_csv_like(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(part).strip() for part in value if str(part).strip()]
    return [part.strip() for part in str(value).split(",") if part.strip()]


def _unique_sorted(values: Iterable[str]) -> List[str]:
    return sorted({value for value in values if value}, key=str)


def _query_id_from_run_dir(run_dir: str | Path) -> str:
    name = Path(run_dir).name
    match = re.search(r"(q\d+)", name)
    return match.group(1) if match else ""


def _safe_read_json(path: Path) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def collect_ranking_anomaly_audit_for_run(run_dir: str | Path, query_id: str | None = None) -> Dict[str, Any]:
    """Collect recoverable ranking-output anomalies from ranked JSON files.

    These are not fatal invalid cases. They are recorded separately so the main
    candidate-ranking Excel stays clean while anomaly details remain auditable.
    """
    run_dir = Path(run_dir)
    qid = str(query_id or _query_id_from_run_dir(run_dir))
    rows: List[Dict[str, Any]] = []

    for mode in MODE_ORDER:
        mode_dir = run_dir / mode
        if not mode_dir.exists():
            continue
        for path in sorted(mode_dir.glob("2_ranked_s*.json")):
            match = re.search(r"2_ranked_s(.+)\.json$", path.name)
            subtask_id = match.group(1) if match else ""
            data = _safe_read_json(path)
            if not isinstance(data, list):
                continue

            anomaly_items = [item for item in data if isinstance(item, dict) and item.get("ranking_anomaly")]
            if not anomaly_items:
                continue

            first = anomaly_items[0]
            rows.append(
                {
                    "Query_ID": qid,
                    "Subtask": subtask_id,
                    "Mode": mode,
                    "Ranking_Anomaly": 1,
                    "Ranking_Anomaly_Reason": first.get("ranking_anomaly_reason", ""),
                    "Missing_API_IDs": ", ".join(_split_csv_like(first.get("missing_api_ids"))),
                    "Unknown_API_IDs": ", ".join(_split_csv_like(first.get("unknown_api_ids"))),
                    "Duplicate_API_IDs": ", ".join(_split_csv_like(first.get("duplicate_api_ids"))),
                    "Expected_API_Count": first.get("expected_api_count", ""),
                    "Actual_API_Count": first.get("actual_api_count", ""),
                    "Returned_API_Count": first.get("returned_api_count", ""),
                }
            )
    return {"ranking_anomaly_rows": rows}


def build_mode_anomaly_rows(
    duplicate_audit: Dict[str, Any] | None,
    hallucination_audit: Dict[str, Any] | None,
    ranking_anomaly_audit: Dict[str, Any] | None = None,
) -> List[Dict[str, Any]]:
    duplicated_by_key: Dict[Tuple[str, str, str], List[str]] = defaultdict(list)
    hallucinated_by_key: Dict[Tuple[str, str, str], List[str]] = defaultdict(list)
    ranking_by_key: Dict[Tuple[str, str, str], Dict[str, Any]] = {}

    if duplicate_audit:
        for row in duplicate_audit.get("duplicate_rows", []):
            key = (
                str(row.get("Query_ID", "")),
                str(row.get("Sub_Task", "")),
                str(row.get("Mode", "")),
            )
            duplicated_by_key[key].append(str(row.get("Selected_API", "")).strip())

        for row in duplicate_audit.get("summary_rows", []):
            key = (
                str(row.get("Query_ID", "")),
                str(row.get("Sub_Task", "")),
                str(row.get("Mode", "")),
            )
            duplicated_by_key[key].extend(
                part.split(" x", 1)[0].strip() for part in _split_csv_like(row.get("Duplicated_APIs", ""))
            )

    if hallucination_audit:
        for row in hallucination_audit.get("mode_detail_rows", []):
            key = (
                str(row.get("Query_ID", "")),
                str(row.get("Sub_Task", "")),
                str(row.get("Mode", "")),
            )
            hallucinated_by_key[key].append(str(row.get("Selected_API", "")).strip())

        for row in hallucination_audit.get("mode_summary_rows", []):
            key = (
                str(row.get("Query_ID", "")),
                str(row.get("Sub_Task", "")),
                str(row.get("Mode", "")),
            )
            hallucinated_by_key[key].extend(_split_csv_like(row.get("Hallucinated_APIs", "")))

    if ranking_anomaly_audit:
        for row in ranking_anomaly_audit.get("ranking_anomaly_rows", []):
            key = (
                str(row.get("Query_ID", "")),
                str(row.get("Subtask", row.get("Sub_Task", ""))),
                str(row.get("Mode", "")),
            )
            ranking_by_key[key] = dict(row)

    keys = sorted(
        set(duplicated_by_key) | set(hallucinated_by_key) | set(ranking_by_key),
        key=lambda key: (_query_sort_key(key[0]), _subtask_sort_key(key[1]), _mode_sort_key(key[2])),
    )

    rows: List[Dict[str, Any]] = []
    for query_id, subtask_id, mode in keys:
        duplicated = _unique_sorted(duplicated_by_key.get((query_id, subtask_id, mode), []))
        hallucinated = _unique_sorted(hallucinated_by_key.get((query_id, subtask_id, mode), []))
        ranking = ranking_by_key.get((query_id, subtask_id, mode), {})
        if not duplicated and not hallucinated and not ranking:
            continue
        rows.append(
            {
                "Query_ID": query_id,
                "Subtask": subtask_id,
                "Mode": mode,
                "APIs_Hallucinated": ", ".join(hallucinated),
                "APIs_Duplicated": ", ".join(duplicated),
                "Ranking_Anomaly": ranking.get("Ranking_Anomaly", ""),
                "Ranking_Anomaly_Reason": ranking.get("Ranking_Anomaly_Reason", ""),
                "Missing_API_IDs": ranking.get("Missing_API_IDs", ""),
                "Unknown_API_IDs": ranking.get("Unknown_API_IDs", ""),
                "Duplicate_API_IDs": ranking.get("Duplicate_API_IDs", ""),
                "Expected_API_Count": ranking.get("Expected_API_Count", ""),
                "Actual_API_Count": ranking.get("Actual_API_Count", ""),
                "Returned_API_Count": ranking.get("Returned_API_Count", ""),
            }
        )
    return rows


def write_mode_anomaly_excel(
    duplicate_audit: Dict[str, Any] | None,
    hallucination_audit: Dict[str, Any] | None,
    out_path: str | Path,
    ranking_anomaly_audit: Dict[str, Any] | None = None,
) -> Path:
    rows = build_mode_anomaly_rows(duplicate_audit, hallucination_audit, ranking_anomaly_audit)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mode Anomalies"

    ws.append(COLUMNS)
    for idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in rows:
        ws.append([row.get(column, "") for column in COLUMNS])
        row_fill = ALERT_FILL if row.get("APIs_Hallucinated") or row.get("Unknown_API_IDs") else WARNING_FILL
        for idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=ws.max_row, column=idx).fill = row_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 14), 70)

    wb.save(out_path)
    return out_path
