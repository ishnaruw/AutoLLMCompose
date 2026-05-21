from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

MODE_ORDER = ["no_qos", "qos_pure_llm", "qos_topsis", "qos_hybrid"]

EVAL_COLUMNS = [
    "Query_ID",
    "Mode",
    "Composition_Validity",
    "Invalid_Reason",
    "Planned_API_Count",
    "Covered_Subtask_Count",
    "Total_Subtask_Count",
    "Composition_Completeness",
    "Composition_Completeness_Gate",
    "Functional_Coverage",
    "Total_Response_Time",
    "Bottleneck_Throughput",
    "Average_Workflow_Availability",
    "Normalized_Response_Time_Score",
    "Normalized_Throughput_Score",
    "Normalized_Availability_Score",
    "Normalized_QoS_Score",
    "QoS_Adjusted_Composition_Score",
    "Planner_Output_File",
]

SUMMARY_COLUMNS = [
    "Mode",
    "Valid",
    "Complete",
    "QoS_Adjusted_Composition_Score",
    "Rank_By_QoS_Adjusted_Score",
    "Functional_Coverage",
    "Normalized_QoS_Score",
    "Total_Response_Time",
    "Bottleneck_Throughput",
    "Average_Workflow_Availability",
    "Composition_Completeness",
    "Composition_Completeness_Gate",
    "Composition_Validity",
    "Short_Interpretation",
]

BEST_MODE_SUMMARY_COLUMNS = [
    "Metric",
    "Direction",
    "Best_Value",
    "Best_Modes",
    "Tie",
]

WORKFLOW_COLUMNS = [
    "Query_ID",
    "Mode",
    "Step",
    "Subtask_ID",
    "API_ID",
    "Functional_Match",
    "rt_ms",
    "tp_rps",
    "availability",
    "Action",
    "Input_From_Previous_Step",
    "Output_To_Next_Step",
    "Why",
]

DEFINITION_ROWS = [
    {
        "Metric": "Composition_Validity",
        "Category": "Planning quality",
        "Definition": "Binary diagnostic metric indicating whether the planner output is structurally usable.",
        "Direction": "Higher is better",
    },
    {
        "Metric": "Composition_Completeness",
        "Category": "Planning quality",
        "Definition": "Fraction of decomposed subtasks covered by valid planned steps.",
        "Direction": "Higher is better",
    },
    {
        "Metric": "Composition_Completeness_Gate",
        "Category": "Planning quality",
        "Definition": "Hard gate set to 1 only when the workflow covers all subtasks, otherwise 0.",
        "Direction": "Higher is better",
    },
    {
        "Metric": "Functional_Coverage",
        "Category": "Functional suitability",
        "Definition": "Fraction of planned APIs with Functional Match Label = 1.",
        "Direction": "Higher is better",
    },
    {
        "Metric": "Total_Response_Time",
        "Category": "QoS quality",
        "Definition": "Sum of response times across planned APIs.",
        "Direction": "Lower is better",
    },
    {
        "Metric": "Bottleneck_Throughput",
        "Category": "QoS quality",
        "Definition": "Minimum throughput across planned APIs.",
        "Direction": "Higher is better",
    },
    {
        "Metric": "Average_Workflow_Availability",
        "Category": "QoS quality",
        "Definition": "Average availability across planned APIs.",
        "Direction": "Higher is better",
    },
    {
        "Metric": "Normalized_QoS_Score",
        "Category": "QoS quality",
        "Definition": "Average of normalized response time, throughput, and availability scores.",
        "Direction": "Higher is better",
    },
    {
        "Metric": "QoS_Adjusted_Composition_Score",
        "Category": "Overall comparison",
        "Definition": "Completeness-gated weighted score using 70% functional coverage and 30% normalized QoS score.",
        "Direction": "Higher is better",
    },
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
BEST_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
INVALID_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
INCOMPLETE_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
COMPLETENESS_GATE_THRESHOLD = 0.999
BEST_MODE_TOLERANCE = 1e-9

BEST_METRIC_SPECS = [
    ("qos_adjusted_composition_score", "QoS_Adjusted_Composition_Score", True, "Higher is better"),
    ("functional_coverage", "Functional_Coverage", True, "Higher is better"),
    ("normalized_qos_score", "Normalized_QoS_Score", True, "Higher is better"),
    ("total_response_time", "Total_Response_Time", False, "Lower is better"),
    ("bottleneck_throughput", "Bottleneck_Throughput", True, "Higher is better"),
    ("average_workflow_availability", "Average_Workflow_Availability", True, "Higher is better"),
    ("composition_completeness", "Composition_Completeness", True, "Higher is better"),
]


def _read_json(path: Path, default: Any) -> Any:
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default
    return default


def _read_json_with_error(path: Path) -> Tuple[Any, str | None]:
    try:
        if not path.exists():
            return None, "missing_file"
        return json.loads(path.read_text(encoding="utf-8")), None
    except json.JSONDecodeError as exc:
        return None, f"invalid_json:{exc.msg}"
    except Exception as exc:
        return None, f"read_error:{type(exc).__name__}"


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def _as_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        return float(value)
    except Exception:
        return None


def _average(values: List[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def _as_int_label(value: Any) -> int:
    if isinstance(value, bool):
        return 1 if value else 0
    if value in (0, 1):
        return int(value)
    if isinstance(value, str):
        text = value.strip().lower()
        if text in {"1", "true", "yes", "relevant", "match"}:
            return 1
    return 0


def _subtask_id(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def _rel(path: Path, root: Path) -> str:
    try:
        return str(path.resolve().relative_to(root.resolve()))
    except Exception:
        return str(path)


def _load_subtasks(query_dir: Path) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    data = _read_json(query_dir / "0_decomposer.json", [])
    if isinstance(data, dict):
        subtasks = data.get("subtasks", [])
    else:
        subtasks = data
    if not isinstance(subtasks, list):
        subtasks = []
    order = {_subtask_id(item.get("id")): idx for idx, item in enumerate(subtasks)}
    return subtasks, order


def _load_selected(query_dir: Path, mode: str) -> Tuple[Dict[str, Dict[str, Any]], List[str]]:
    selected: Dict[str, Dict[str, Any]] = {}
    missing_files: List[str] = []
    mode_dir = query_dir / mode
    files = sorted(mode_dir.glob("3_selected_s*.json"))
    if not files:
        missing_files.append(f"missing_selected_api_files:{_rel(mode_dir / '3_selected_s*.json', query_dir)}")
    for path in files:
        rows, error = _read_json_with_error(path)
        if error:
            missing_files.append(f"invalid_selected_api_file:{_rel(path, query_dir)}:{error}")
            continue
        if not isinstance(rows, list):
            missing_files.append(f"invalid_selected_api_file:{_rel(path, query_dir)}:not_list")
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            api_id = str(row.get("api_id") or row.get("Selected_API") or "").strip()
            if api_id:
                selected.setdefault(api_id, row)
    return selected, missing_files


def _candidate_lookup(rows_path: Path) -> Dict[Tuple[str, str, str], Dict[str, Any]]:
    lookup: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    rows = _read_json(rows_path, [])
    if not isinstance(rows, list):
        return lookup
    for row in rows:
        if not isinstance(row, dict):
            continue
        mode = str(row.get("Mode") or row.get("mode") or "").strip()
        sid = _subtask_id(row.get("Sub Task", row.get("subtask_id")))
        api_id = str(row.get("Selected_API") or row.get("api_id") or "").strip()
        if mode and sid and api_id:
            lookup[(mode, sid, api_id)] = row
    return lookup


def _qos_from_selected(row: Dict[str, Any] | None) -> Tuple[float | None, float | None, float | None]:
    if not isinstance(row, dict):
        return None, None, None
    qos = row.get("qos")
    service = row.get("service")
    if not isinstance(qos, dict) and isinstance(service, dict):
        qos = service.get("qos")
    if not isinstance(qos, dict):
        qos = {}
    return _as_float(qos.get("rt_ms")), _as_float(qos.get("tp_rps")), _as_float(qos.get("availability"))


def _metrics_for_step(
    *,
    candidate_rows: Dict[Tuple[str, str, str], Dict[str, Any]],
    selected_rows: Dict[str, Dict[str, Any]],
    mode: str,
    subtask_id: str,
    api_id: str,
) -> Tuple[int, float | None, float | None, float | None]:
    candidate = candidate_rows.get((mode, subtask_id, api_id))
    functional = _as_int_label(candidate.get("Functional Match (0/1)")) if isinstance(candidate, dict) else 0
    rt = _as_float(candidate.get("QoS_RT")) if isinstance(candidate, dict) else None
    tp = _as_float(candidate.get("QoS_TP")) if isinstance(candidate, dict) else None
    av = _as_float(candidate.get("QoS Availability")) if isinstance(candidate, dict) else None
    fallback_rt, fallback_tp, fallback_av = _qos_from_selected(selected_rows.get(api_id))
    return functional, rt if rt is not None else fallback_rt, tp if tp is not None else fallback_tp, av if av is not None else fallback_av


def _required_step_fields_present(step: Dict[str, Any]) -> bool:
    required_present = [
        "step",
        "api_id",
        "subtask_id",
        "action",
        "input_from_previous_step",
        "output_to_next_step",
        "why",
    ]
    required_non_null = ["step", "api_id", "subtask_id", "action", "why"]
    return all(key in step for key in required_present) and all(step.get(key) is not None for key in required_non_null)


def _evaluate_mode(
    *,
    query_dir: Path,
    query_id: str,
    mode: str,
    subtasks: List[Dict[str, Any]],
    subtask_order: Dict[str, int],
    candidate_rows: Dict[Tuple[str, str, str], Dict[str, Any]],
) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    selected_rows, selected_issues = _load_selected(query_dir, mode)
    selected_ids = set(selected_rows.keys())
    planner_path = query_dir / mode / "4_planner.json"
    planner, planner_error = _read_json_with_error(planner_path)
    workflow_rows: List[Dict[str, Any]] = []
    reasons: List[str] = list(selected_issues)

    if planner_error == "missing_file":
        reasons.append("missing_planner_output")
    elif planner_error:
        reasons.append(f"invalid_planner_output:{planner_error}")

    primary_plan = planner.get("primary_plan") if isinstance(planner, dict) else None
    if not isinstance(primary_plan, dict):
        reasons.append("missing_or_invalid_primary_plan")
        steps = []
    else:
        steps = primary_plan.get("steps")
        if not isinstance(steps, list):
            reasons.append("missing_or_invalid_steps")
            steps = []

    if not steps:
        reasons.append("empty_plan_steps")
    if not selected_ids:
        reasons.append("empty_selected_api_list")

    total_subtasks = len(subtasks)
    covered_valid_subtasks: set[str] = set()
    planned_api_count = len(steps)
    functional_matches = 0
    rt_values: List[float] = []
    tp_values: List[float] = []
    av_values: List[float] = []
    has_missing_qos = False
    previous_order = -1

    for idx, raw_step in enumerate(steps, start=1):
        step = raw_step if isinstance(raw_step, dict) else {}
        if not step or not _required_step_fields_present(step):
            reasons.append(f"step_{idx}_missing_required_fields")
        api_id = str(step.get("api_id") or "").strip()
        sid = _subtask_id(step.get("subtask_id"))
        if api_id not in selected_ids:
            reasons.append(f"step_{idx}_api_not_selected:{api_id or '<missing>'}")
        if sid not in subtask_order:
            reasons.append(f"step_{idx}_invalid_subtask_id:{sid or '<missing>'}")
        else:
            current_order = subtask_order[sid]
            if current_order < previous_order:
                reasons.append(f"step_{idx}_subtask_order_violation")
            previous_order = max(previous_order, current_order)

        functional, rt, tp, av = _metrics_for_step(
            candidate_rows=candidate_rows,
            selected_rows=selected_rows,
            mode=mode,
            subtask_id=sid,
            api_id=api_id,
        )
        functional_matches += functional
        if rt is None or tp is None or av is None:
            has_missing_qos = True
        else:
            rt_values.append(rt)
            tp_values.append(tp)
            av_values.append(av)

        if api_id in selected_ids and sid in subtask_order and _required_step_fields_present(step):
            covered_valid_subtasks.add(sid)

        workflow_rows.append(
            {
                "Query_ID": query_id,
                "Mode": mode,
                "Step": step.get("step", idx),
                "Subtask_ID": sid,
                "API_ID": api_id,
                "Functional_Match": functional,
                "rt_ms": rt,
                "tp_rps": tp,
                "availability": av,
                "Action": step.get("action", ""),
                "Input_From_Previous_Step": step.get("input_from_previous_step", ""),
                "Output_To_Next_Step": step.get("output_to_next_step", ""),
                "Why": step.get("why", ""),
            }
        )

    seen_reasons: List[str] = []
    for reason in reasons:
        if reason and reason not in seen_reasons:
            seen_reasons.append(reason)

    valid = 1 if not seen_reasons else 0
    completeness = (len(covered_valid_subtasks) / total_subtasks) if total_subtasks else 0.0
    completeness_gate = 1.0 if completeness >= COMPLETENESS_GATE_THRESHOLD else 0.0
    qos_available = valid == 1 and planned_api_count > 0 and not has_missing_qos and len(rt_values) == planned_api_count
    row = {
        "Query_ID": query_id,
        "Mode": mode,
        "Composition_Validity": valid,
        "Invalid_Reason": "; ".join(seen_reasons),
        "Planned_API_Count": planned_api_count,
        "Covered_Subtask_Count": len(covered_valid_subtasks),
        "Total_Subtask_Count": total_subtasks,
        "Composition_Completeness": completeness,
        "Composition_Completeness_Gate": completeness_gate,
        "Functional_Coverage": (functional_matches / planned_api_count) if planned_api_count else 0.0,
        "Total_Response_Time": round(sum(rt_values), 6) if qos_available else None,
        "Bottleneck_Throughput": round(min(tp_values), 6) if qos_available else None,
        "Average_Workflow_Availability": round(_average(av_values), 6) if qos_available else None,
        "Normalized_Response_Time_Score": 0.0 if valid == 0 or has_missing_qos else None,
        "Normalized_Throughput_Score": 0.0 if valid == 0 or has_missing_qos else None,
        "Normalized_Availability_Score": 0.0 if valid == 0 or has_missing_qos else None,
        "Normalized_QoS_Score": 0.0 if valid == 0 or has_missing_qos else None,
        "QoS_Adjusted_Composition_Score": None,
        "Planner_Output_File": _rel(planner_path, query_dir),
    }
    return row, workflow_rows


def _normalize(rows: List[Dict[str, Any]]) -> None:
    metric_specs = [
        ("Total_Response_Time", "Normalized_Response_Time_Score", False),
        ("Bottleneck_Throughput", "Normalized_Throughput_Score", True),
        ("Average_Workflow_Availability", "Normalized_Availability_Score", True),
    ]
    for raw_key, norm_key, higher_better in metric_specs:
        values = [_as_float(row.get(raw_key)) for row in rows if _as_float(row.get(raw_key)) is not None]
        if not values:
            for row in rows:
                if row.get(norm_key) is None:
                    row[norm_key] = 0.0
            continue
        min_value = min(values)
        max_value = max(values)
        for row in rows:
            raw = _as_float(row.get(raw_key))
            if raw is None:
                row[norm_key] = 0.0
            elif max_value == min_value:
                row[norm_key] = 1.0
            elif higher_better:
                row[norm_key] = round((raw - min_value) / (max_value - min_value), 6)
            else:
                row[norm_key] = round((max_value - raw) / (max_value - min_value), 6)

    for row in rows:
        if row.get("Composition_Validity") == 0 or any(row.get(k) is None for k in ["Total_Response_Time", "Bottleneck_Throughput", "Average_Workflow_Availability"]):
            row["Normalized_QoS_Score"] = 0.0
        else:
            row["Normalized_QoS_Score"] = round(
                (
                    float(row["Normalized_Response_Time_Score"])
                    + float(row["Normalized_Throughput_Score"])
                    + float(row["Normalized_Availability_Score"])
                )
                / 3.0,
                6,
            )
        validity = float(row.get("Composition_Validity") or 0.0)
        completeness_gate = float(row.get("Composition_Completeness_Gate") or 0.0)
        # The final score uses completeness as a hard gate because incomplete workflows do not fully
        # satisfy the user request. Among complete workflows, functional coverage is weighted higher
        # than QoS because functional suitability is required before QoS optimization is meaningful.
        if validity <= 0.0 or completeness_gate <= 0.0:
            row["QoS_Adjusted_Composition_Score"] = 0.0
        else:
            row["QoS_Adjusted_Composition_Score"] = round(
                0.7 * float(row.get("Functional_Coverage") or 0.0)
                + 0.3 * float(row.get("Normalized_QoS_Score") or 0.0),
                6,
            )


def _rank(rows: List[Dict[str, Any]], key: str, *, reverse: bool) -> Dict[str, int | None]:
    available = [(row["Mode"], _as_float(row.get(key))) for row in rows if _as_float(row.get(key)) is not None]
    available.sort(key=lambda item: item[1], reverse=reverse)
    ranks: Dict[str, int | None] = {row["Mode"]: None for row in rows}
    last_value: float | None = None
    last_rank = 0
    for idx, (mode, value) in enumerate(available, start=1):
        if last_value is None or value != last_value:
            last_rank = idx
            last_value = value
        ranks[mode] = last_rank
    return ranks


def get_best_modes(rows: List[Dict[str, Any]], metric: str, higher_is_better: bool = True, tolerance: float = BEST_MODE_TOLERANCE) -> Dict[str, Any]:
    values: List[Tuple[str, float]] = []
    mode_order = {mode: idx for idx, mode in enumerate(MODE_ORDER)}
    for row in rows:
        value = _as_float(row.get(metric))
        if value is not None:
            values.append((str(row.get("Mode") or ""), value))
    if not values:
        return {"best_value": None, "best_modes": [], "is_tie": False}

    best_value = max(value for _, value in values) if higher_is_better else min(value for _, value in values)
    best_modes = [
        mode
        for mode, value in values
        if abs(value - best_value) <= tolerance
    ]
    best_modes = sorted(best_modes, key=lambda mode: mode_order.get(mode, len(mode_order)))
    return {
        "best_value": best_value,
        "best_modes": best_modes,
        "is_tie": len(best_modes) > 1,
    }


def _summary_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    score_ranks = _rank(rows, "QoS_Adjusted_Composition_Score", reverse=True)
    best_score = min((rank for rank in score_ranks.values() if rank is not None), default=None)
    best_score_modes = set(get_best_modes(rows, "QoS_Adjusted_Composition_Score", higher_is_better=True)["best_modes"])
    score_tied = len(best_score_modes) > 1

    out: List[Dict[str, Any]] = []
    for row in rows:
        mode = row["Mode"]
        complete = float(row.get("Composition_Completeness_Gate") or 0.0) >= 1.0
        functional = float(row.get("Functional_Coverage") or 0.0)
        normalized_qos = float(row.get("Normalized_QoS_Score") or 0.0)
        if row.get("Composition_Validity") == 0:
            interpretation = "Invalid composition plan"
        elif not complete:
            interpretation = "Incomplete workflow, final score gated to 0"
        elif mode in best_score_modes or score_ranks.get(mode) == best_score:
            interpretation = "Tied best overall composition score" if score_tied else "Best overall composition score"
        elif normalized_qos >= 0.667 and functional < 0.9:
            interpretation = "Strong QoS but weaker functional coverage"
        elif functional >= 0.9 and normalized_qos < 0.667:
            interpretation = "Strong functional coverage with moderate QoS"
        elif functional >= 0.9:
            interpretation = "Complete and functionally strong workflow"
        else:
            interpretation = "Complete workflow with mixed functional and QoS tradeoffs"
        out.append(
            {
                "Mode": mode,
                "Valid": "Yes" if row.get("Composition_Validity") == 1 else "No",
                "Complete": "Yes" if complete else "No",
                "QoS_Adjusted_Composition_Score": row.get("QoS_Adjusted_Composition_Score"),
                "Rank_By_QoS_Adjusted_Score": score_ranks.get(mode),
                "Functional_Coverage": row.get("Functional_Coverage"),
                "Normalized_QoS_Score": row.get("Normalized_QoS_Score"),
                "Total_Response_Time": row.get("Total_Response_Time"),
                "Bottleneck_Throughput": row.get("Bottleneck_Throughput"),
                "Average_Workflow_Availability": row.get("Average_Workflow_Availability"),
                "Composition_Completeness": row.get("Composition_Completeness"),
                "Composition_Completeness_Gate": row.get("Composition_Completeness_Gate"),
                "Composition_Validity": row.get("Composition_Validity"),
                "Short_Interpretation": interpretation,
            }
        )
    return out


def _validity_summary(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    by_mode: Dict[str, Dict[str, Any]] = {}
    invalid_modes: List[str] = []
    invalid_reason_counts: Dict[str, int] = {}
    for row in rows:
        mode = str(row.get("Mode") or "")
        valid = int(row.get("Composition_Validity") or 0)
        reasons = [part.strip() for part in str(row.get("Invalid_Reason") or "").split(";") if part.strip()]
        if not valid:
            invalid_modes.append(mode)
            for reason in reasons:
                invalid_reason_counts[reason] = invalid_reason_counts.get(reason, 0) + 1
        by_mode[mode] = {
            "composition_validity": valid,
            "invalid_reasons": reasons,
            "planned_api_count": row.get("Planned_API_Count"),
            "covered_subtask_count": row.get("Covered_Subtask_Count"),
            "total_subtask_count": row.get("Total_Subtask_Count"),
            "composition_completeness": row.get("Composition_Completeness"),
            "composition_completeness_gate": row.get("Composition_Completeness_Gate"),
            "functional_coverage": row.get("Functional_Coverage"),
        }
    return {
        "valid_mode_count": sum(1 for row in rows if int(row.get("Composition_Validity") or 0) == 1),
        "complete_mode_count": sum(1 for row in rows if float(row.get("Composition_Completeness_Gate") or 0.0) >= 1.0),
        "invalid_mode_count": len(invalid_modes),
        "invalid_modes": invalid_modes,
        "invalid_reason_counts": invalid_reason_counts,
        "by_mode": by_mode,
    }


def _best_mode_by(rows: List[Dict[str, Any]], key: str, *, higher_better: bool) -> Dict[str, Any] | None:
    best = get_best_modes(rows, key, higher_is_better=higher_better)
    if not best["best_modes"]:
        return None
    return {
        "mode": best["best_modes"][0],
        "value": best["best_value"],
        "compatibility_note": "First best mode only; use best_modes_by_* fields for tie-aware reporting.",
    }


def _score_summary(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    summary: Dict[str, Any] = {
        "valid_mode_count": sum(1 for row in rows if int(row.get("Composition_Validity") or 0) == 1),
        "complete_mode_count": sum(1 for row in rows if float(row.get("Composition_Completeness_Gate") or 0.0) >= 1.0),
        "evaluated_modes": MODE_ORDER,
        "scoring_formula": (
            "QoS_Adjusted_Composition_Score = 0 if invalid or incomplete; otherwise "
            "0.7 * Functional_Coverage + 0.3 * Normalized_QoS_Score"
        ),
    }
    for slug, metric, higher_better, _ in BEST_METRIC_SPECS:
        best = get_best_modes(rows, metric, higher_is_better=higher_better)
        summary[f"best_modes_by_{slug}"] = best["best_modes"]
        summary[f"best_{slug}"] = best["best_value"]
        summary[f"is_{slug}_tie"] = best["is_tie"]
        summary[f"best_mode_by_{slug}"] = _best_mode_by(rows, metric, higher_better=higher_better)
    return summary


def _best_mode_summary_rows(score_summary: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for slug, metric, _, direction in BEST_METRIC_SPECS:
        best_modes = score_summary.get(f"best_modes_by_{slug}") or []
        rows.append(
            {
                "Metric": metric,
                "Direction": direction,
                "Best_Value": score_summary.get(f"best_{slug}"),
                "Best_Modes": ", ".join(best_modes) if best_modes else "",
                "Tie": "Yes" if score_summary.get(f"is_{slug}_tie") else "No",
            }
        )
    return rows


def _issue_log_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    issue_rows: List[Dict[str, Any]] = []
    for row in rows:
        mode = str(row.get("Mode") or "")
        reasons = [part.strip() for part in str(row.get("Invalid_Reason") or "").split(";") if part.strip()]
        for reason in reasons:
            issue_rows.append(
                {
                    "Query_ID": row.get("Query_ID"),
                    "Mode": mode,
                    "Composition_Validity": row.get("Composition_Validity"),
                    "Issue": reason,
                    "Planner_Output_File": row.get("Planner_Output_File"),
                    "Planned_API_Count": row.get("Planned_API_Count"),
                    "Covered_Subtask_Count": row.get("Covered_Subtask_Count"),
                    "Total_Subtask_Count": row.get("Total_Subtask_Count"),
                }
            )
    return issue_rows


def _write_issue_text_log(path: Path, issue_rows: List[Dict[str, Any]], *, query_id: str) -> None:
    lines = [f"Composition validity issue log for query {query_id}"]
    if not issue_rows:
        lines.append("No composition validity issues detected.")
    else:
        for row in issue_rows:
            lines.append(
                " | ".join(
                    [
                        f"mode={row.get('Mode')}",
                        f"validity={row.get('Composition_Validity')}",
                        f"issue={row.get('Issue')}",
                        f"planner={row.get('Planner_Output_File')}",
                    ]
                )
            )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _style_header(ws, columns: List[str]) -> None:
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _finalize_sheet(ws, numeric_columns: set[str] | None = None) -> None:
    numeric_columns = numeric_columns or set()
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header in numeric_columns and cell.row > 1 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.000"
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 60)


def _append_rows(ws, columns: List[str], rows: List[Dict[str, Any]], numeric_columns: set[str] | None = None) -> None:
    _style_header(ws, columns)
    for row in rows:
        ws.append([row.get(column) for column in columns])
    _finalize_sheet(ws, numeric_columns)


def _highlight_main_sheet(ws, rows: List[Dict[str, Any]]) -> None:
    best_specs = [
        ("Total_Response_Time", False),
        ("Bottleneck_Throughput", True),
        ("Average_Workflow_Availability", True),
        ("Functional_Coverage", True),
        ("Normalized_QoS_Score", True),
        ("QoS_Adjusted_Composition_Score", True),
    ]
    best_values: Dict[str, float] = {}
    for key, higher_better in best_specs:
        values = [_as_float(row.get(key)) for row in rows if _as_float(row.get(key)) is not None]
        if values:
            best_values[key] = max(values) if higher_better else min(values)

    header_to_col = {str(cell.value): cell.column for cell in ws[1]}
    for row_idx, row in enumerate(rows, start=2):
        if row.get("Composition_Validity") == 0:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = INVALID_FILL
        elif float(row.get("Composition_Completeness_Gate") or 0.0) < 1.0:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = INCOMPLETE_FILL
        for key, best in best_values.items():
            value = _as_float(row.get(key))
            if value is not None and abs(value - best) <= BEST_MODE_TOLERANCE and key in header_to_col:
                ws.cell(row=row_idx, column=header_to_col[key]).fill = BEST_FILL


def _write_excel(path: Path, rows: List[Dict[str, Any]], summary: List[Dict[str, Any]], workflow: List[Dict[str, Any]], score_summary: Dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Composition_Evaluation"
    numeric = {
        "Composition_Completeness",
        "Composition_Completeness_Gate",
        "Functional_Coverage",
        "Total_Response_Time",
        "Bottleneck_Throughput",
        "Average_Workflow_Availability",
        "Normalized_Response_Time_Score",
        "Normalized_Throughput_Score",
        "Normalized_Availability_Score",
        "Normalized_QoS_Score",
        "QoS_Adjusted_Composition_Score",
        "Best_Value",
    }
    _append_rows(ws, EVAL_COLUMNS, rows, numeric)
    _highlight_main_sheet(ws, rows)

    ws = wb.create_sheet("Mode_Summary")
    _append_rows(ws, SUMMARY_COLUMNS, summary, numeric)

    ws = wb.create_sheet("Best_Mode_Summary")
    _append_rows(ws, BEST_MODE_SUMMARY_COLUMNS, _best_mode_summary_rows(score_summary), numeric)

    ws = wb.create_sheet("Planned_Workflow")
    _append_rows(ws, WORKFLOW_COLUMNS, workflow, {"rt_ms", "tp_rps", "availability"})

    ws = wb.create_sheet("Metric_Definitions")
    _append_rows(ws, ["Metric", "Category", "Definition", "Direction"], DEFINITION_ROWS)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def evaluate_composition_qos(*, query_dir: Path, query_id: str | None = None, output_dir: Path | None = None) -> Dict[str, Any]:
    query_dir = Path(query_dir)
    meta = _read_json(query_dir / "meta.json", {})
    query_id = str(query_id or meta.get("query_id") or query_dir.name)
    output_dir = Path(output_dir or query_dir / "evaluation")
    output_dir.mkdir(parents=True, exist_ok=True)

    subtasks, subtask_order = _load_subtasks(query_dir)
    candidate_rows_path = output_dir / f"query_{query_id}_candidate_api_rankings_rows.json"
    candidate_rows = _candidate_lookup(candidate_rows_path)

    rows: List[Dict[str, Any]] = []
    workflow_rows: List[Dict[str, Any]] = []
    for mode in MODE_ORDER:
        row, mode_workflow = _evaluate_mode(
            query_dir=query_dir,
            query_id=query_id,
            mode=mode,
            subtasks=subtasks,
            subtask_order=subtask_order,
            candidate_rows=candidate_rows,
        )
        rows.append(row)
        workflow_rows.extend(mode_workflow)

    _normalize(rows)
    summary_rows = _summary_rows(rows)
    validity_summary = _validity_summary(rows)
    score_summary = _score_summary(rows)
    issue_rows = _issue_log_rows(rows)

    rows_path = output_dir / f"query_{query_id}_composition_qos_eval_rows.json"
    summary_path = output_dir / f"query_{query_id}_composition_qos_eval_summary.json"
    xlsx_path = output_dir / f"query_{query_id}_composition_qos_eval.xlsx"
    issues_json_path = output_dir / f"query_{query_id}_composition_validity_issues.json"
    issues_log_path = output_dir / f"query_{query_id}_composition_validity_issues.log"

    _write_json(rows_path, rows)
    _write_json(issues_json_path, issue_rows)
    _write_issue_text_log(issues_log_path, issue_rows, query_id=query_id)
    summary_payload = {
        "query_id": query_id,
        "rows_json": str(rows_path),
        "excel": str(xlsx_path),
        "composition_validity_issues_json": str(issues_json_path),
        "composition_validity_issues_log": str(issues_log_path),
        "candidate_api_rankings_rows_json": str(candidate_rows_path),
        "modes": MODE_ORDER,
        "evaluated_modes": MODE_ORDER,
        "scoring_formula": score_summary["scoring_formula"],
        "valid_mode_count": score_summary["valid_mode_count"],
        "complete_mode_count": score_summary["complete_mode_count"],
        "composition_validity_summary": validity_summary,
        "score_summary": score_summary,
        "summary_rows": summary_rows,
    }
    for slug, _, _, _ in BEST_METRIC_SPECS:
        for prefix in ("best_modes_by", "best", "is", "best_mode_by"):
            key = f"{prefix}_{slug}" if prefix != "is" else f"is_{slug}_tie"
            if key in score_summary:
                summary_payload[key] = score_summary[key]
    _write_json(summary_path, summary_payload)
    _write_excel(xlsx_path, rows, summary_rows, workflow_rows, score_summary)

    return {
        "rows_json": rows_path,
        "summary_json": summary_path,
        "excel": xlsx_path,
        "composition_validity_issues_json": issues_json_path,
        "composition_validity_issues_log": issues_log_path,
        "rows": rows,
        "summary_rows": summary_rows,
        "composition_validity_summary": validity_summary,
        "score_summary": score_summary,
    }


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Evaluate deterministic composition-level QoS for one AutoLLMCompose query run.")
    parser.add_argument("query_dir", type=Path)
    parser.add_argument("--query-id")
    parser.add_argument("--output-dir", type=Path)
    args = parser.parse_args()
    result = evaluate_composition_qos(query_dir=args.query_dir, query_id=args.query_id, output_dir=args.output_dir)
    print(json.dumps({k: str(v) for k, v in result.items() if k.endswith("_json") or k == "excel"}, indent=2))
