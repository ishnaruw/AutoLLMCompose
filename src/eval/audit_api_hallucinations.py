from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from src.tools.fetch_services import catalog_path

MODE_ORDER = ["no_qos", "qos_pure_llm", "qos_topsis", "qos_hybrid"]
MODE_INDEX = {name: idx for idx, name in enumerate(MODE_ORDER)}

RETRIEVAL_SUMMARY_COLUMNS = [
    "Provider_Dir",
    "Run_Dir",
    "Query_ID",
    "Sub_Task",
    "Subtask_Purpose",
    "Retrieved_Count",
    "Unique_Retrieved_Count",
    "Retrieved_Not_In_Catalog_Count",
    "Has_Retrieval_Hallucination",
    "Missing_API_IDs",
]

RETRIEVAL_DETAIL_COLUMNS = [
    "Provider_Dir",
    "Run_Dir",
    "Query_ID",
    "Sub_Task",
    "Subtask_Purpose",
    "Retrieved_Rank",
    "API_ID",
    "Issue",
]

MODE_SUMMARY_COLUMNS = [
    "Provider_Dir",
    "Run_Dir",
    "Query_ID",
    "Sub_Task",
    "Mode",
    "Subtask_Purpose",
    "Retrieved_Candidate_Count",
    "Selected_Count",
    "Hallucination_Row_Count",
    "Selected_Not_In_Retrieved_Count",
    "Selected_Not_In_Catalog_Count",
    "Selected_From_Retrieved_But_Catalog_Missing_Count",
    "Has_Mode_Hallucination",
    "Hallucinated_APIs",
]

MODE_DETAIL_COLUMNS = [
    "Provider_Dir",
    "Run_Dir",
    "Query_ID",
    "Sub_Task",
    "Mode",
    "Subtask_Purpose",
    "Selected_API",
    "Mode_Rank",
    "Retrieved_Rank",
    "In_Retrieved_Set",
    "In_Catalog",
    "Issue",
    "Comments",
]

OVERVIEW_COLUMNS = [
    "Key",
    "Groups_Checked",
    "Groups_With_Hallucinations",
    "Hallucinated_API_Count",
]


def _query_sort_key(value: str) -> Tuple[int, str]:
    match = re.search(r"q(\d+)", str(value), flags=re.IGNORECASE)
    if match:
        return int(match.group(1)), str(value)
    return 9999, str(value)


def _subtask_sort_key(value: Any) -> Tuple[int, str]:
    text = str(value)
    return (int(text), text) if text.isdigit() else (9999, text)


def _mode_sort_key(value: str) -> Tuple[int, str]:
    text = str(value)
    return MODE_INDEX.get(text, 9999), text


def _safe_load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _is_query_run_dir(path: Path) -> bool:
    return path.is_dir() and (path / "0_decomposer.json").exists()


def _iter_run_dirs(root_dir: Path) -> Iterable[Path]:
    if _is_query_run_dir(root_dir):
        yield root_dir
        return
    pattern = re.compile(r"^q\d+_\d{8}T\d{6}$", flags=re.IGNORECASE)
    run_dirs = [path for path in root_dir.rglob("*") if path.is_dir() and pattern.match(path.name)]
    run_dirs.sort(key=lambda p: (str(p.parent), _query_sort_key(p.name)))
    for path in run_dirs:
        yield path


def _provider_dir_label(root_dir: Path, run_dir: Path) -> str:
    if root_dir == run_dir:
        return "."
    rel_parent = run_dir.parent.relative_to(root_dir)
    return rel_parent.as_posix() if rel_parent.parts else run_dir.parent.name


def _find_evaluation_dir(run_dir: Path) -> Path | None:
    for candidate in ("evaluation", "functional_match_eval"):
        eval_dir = run_dir / candidate
        if eval_dir.exists():
            return eval_dir
    return None


def _find_rows_json(run_dir: Path) -> Path | None:
    eval_dir = _find_evaluation_dir(run_dir)
    if eval_dir is None:
        return None
    matches = sorted(eval_dir.glob("query_*_candidate_api_rankings_rows.json"))
    return matches[0] if matches else None


def _load_catalog_ids(catalog_path: Path) -> set[str]:
    ids: set[str] = set()
    with catalog_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception:
                continue
            api_id = obj.get("api_id")
            if api_id:
                ids.add(str(api_id))
    return ids


def _load_subtask_map(run_dir: Path) -> Dict[str, str]:
    path = run_dir / "0_decomposer.json"
    if not path.exists():
        return {}
    data = _safe_load_json(path)
    if not isinstance(data, list):
        return {}
    out: Dict[str, str] = {}
    for row in data:
        if not isinstance(row, dict):
            continue
        sub_id = str(row.get("id", "")).strip()
        desc = str(row.get("description", "")).strip()
        if sub_id:
            out[sub_id] = desc
    return out


def _load_retrieved_candidates(run_dir: Path) -> Dict[str, List[Dict[str, Any]]]:
    out: Dict[str, List[Dict[str, Any]]] = {}
    for path in sorted(run_dir.glob("1_retriever_s*.json")):
        match = re.search(r"1_retriever_s(\d+)\.json$", path.name)
        if not match:
            continue
        subtask_id = match.group(1)
        data = _safe_load_json(path)
        if isinstance(data, list):
            out[subtask_id] = data
    return out


def _collect_hallucination_audit_for_runs(root_dir: Path, catalog_path: Path, run_dirs: Iterable[Path]) -> Dict[str, Any]:
    catalog_ids = _load_catalog_ids(catalog_path)
    retrieval_summary_rows: List[Dict[str, Any]] = []
    retrieval_detail_rows: List[Dict[str, Any]] = []
    mode_summary_rows: List[Dict[str, Any]] = []
    mode_detail_rows: List[Dict[str, Any]] = []
    missing_runs: List[str] = []

    for run_dir in run_dirs:
        provider_dir = _provider_dir_label(root_dir, run_dir)
        subtask_map = _load_subtask_map(run_dir)
        retrieved_by_subtask = _load_retrieved_candidates(run_dir)
        rows_json_path = _find_rows_json(run_dir)
        if rows_json_path is None:
            missing_runs.append(run_dir.name)
            continue

        rows_data = _safe_load_json(rows_json_path)
        if not isinstance(rows_data, list):
            missing_runs.append(run_dir.name)
            continue

        retrieved_id_sets: Dict[str, set[str]] = {}
        retrieved_catalog_missing_by_subtask: Dict[str, set[str]] = {}

        for subtask_id, rows in sorted(retrieved_by_subtask.items(), key=lambda item: _subtask_sort_key(item[0])):
            api_ids = [str(row.get("api_id", "")).strip() for row in rows if str(row.get("api_id", "")).strip()]
            unique_ids = set(api_ids)
            missing_api_ids = sorted(api_id for api_id in unique_ids if api_id not in catalog_ids)
            retrieved_id_sets[subtask_id] = unique_ids
            retrieved_catalog_missing_by_subtask[subtask_id] = set(missing_api_ids)

            retrieval_summary_rows.append(
                {
                    "Provider_Dir": provider_dir,
                    "Run_Dir": run_dir.name,
                    "Query_ID": run_dir.name.split("_", 1)[0],
                    "Sub_Task": subtask_id,
                    "Subtask_Purpose": subtask_map.get(subtask_id, ""),
                    "Retrieved_Count": len(rows),
                    "Unique_Retrieved_Count": len(unique_ids),
                    "Retrieved_Not_In_Catalog_Count": len(missing_api_ids),
                    "Has_Retrieval_Hallucination": "Y" if missing_api_ids else "N",
                    "Missing_API_IDs": ", ".join(missing_api_ids),
                }
            )

            if missing_api_ids:
                missing_set = set(missing_api_ids)
                for row in rows:
                    api_id = str(row.get("api_id", "")).strip()
                    if api_id in missing_set:
                        retrieval_detail_rows.append(
                            {
                                "Provider_Dir": provider_dir,
                                "Run_Dir": run_dir.name,
                                "Query_ID": run_dir.name.split("_", 1)[0],
                                "Sub_Task": subtask_id,
                                "Subtask_Purpose": subtask_map.get(subtask_id, ""),
                                "Retrieved_Rank": row.get("retrieved_rank", ""),
                                "API_ID": api_id,
                                "Issue": "retrieved_not_in_catalog",
                            }
                        )

        grouped_mode_rows: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
        for row in rows_data:
            query_id = str(row.get("Query_ID", "")).strip()
            subtask_id = str(row.get("Sub Task", "")).strip()
            mode = str(row.get("Mode", "")).strip()
            grouped_mode_rows[(query_id, subtask_id, mode)].append(row)

        for (query_id, subtask_id, mode), group_rows in sorted(
            grouped_mode_rows.items(),
            key=lambda item: (_query_sort_key(item[0][0]), _subtask_sort_key(item[0][1]), _mode_sort_key(item[0][2])),
        ):
            retrieved_set = retrieved_id_sets.get(subtask_id, set())
            retrieval_catalog_missing = retrieved_catalog_missing_by_subtask.get(subtask_id, set())
            hallucinated_ids: List[str] = []
            selected_not_in_retrieved = 0
            selected_not_in_catalog = 0
            selected_from_retrieved_but_catalog_missing = 0

            for row in group_rows:
                api_id = str(row.get("Selected_API", "")).strip()
                if not api_id:
                    continue

                in_retrieved = api_id in retrieved_set
                in_catalog = api_id in catalog_ids
                issue_flags: List[str] = []
                if not in_retrieved:
                    selected_not_in_retrieved += 1
                    issue_flags.append("selected_not_in_retrieved")
                if not in_catalog:
                    selected_not_in_catalog += 1
                    issue_flags.append("selected_not_in_catalog")
                if in_retrieved and api_id in retrieval_catalog_missing:
                    selected_from_retrieved_but_catalog_missing += 1
                    issue_flags.append("selected_from_retrieved_but_catalog_missing")

                if issue_flags:
                    hallucinated_ids.append(api_id)
                    mode_detail_rows.append(
                        {
                            "Provider_Dir": provider_dir,
                            "Run_Dir": run_dir.name,
                            "Query_ID": query_id,
                            "Sub_Task": subtask_id,
                            "Mode": mode,
                            "Subtask_Purpose": str(row.get("Subtask_Purpose", "")),
                            "Selected_API": api_id,
                            "Mode_Rank": row.get("Mode Rank", ""),
                            "Retrieved_Rank": row.get("Retrieved Rank", ""),
                            "In_Retrieved_Set": "Y" if in_retrieved else "N",
                            "In_Catalog": "Y" if in_catalog else "N",
                            "Issue": ", ".join(issue_flags),
                            "Comments": row.get("Comments", ""),
                        }
                    )

            unique_hallucinated_ids = sorted(set(hallucinated_ids))
            mode_summary_rows.append(
                {
                    "Provider_Dir": provider_dir,
                    "Run_Dir": run_dir.name,
                    "Query_ID": query_id,
                    "Sub_Task": subtask_id,
                    "Mode": mode,
                    "Subtask_Purpose": str(group_rows[0].get("Subtask_Purpose", "")) if group_rows else subtask_map.get(subtask_id, ""),
                    "Retrieved_Candidate_Count": len(retrieved_set),
                    "Selected_Count": len(group_rows),
                    "Hallucination_Row_Count": len(hallucinated_ids),
                    "Selected_Not_In_Retrieved_Count": selected_not_in_retrieved,
                    "Selected_Not_In_Catalog_Count": selected_not_in_catalog,
                    "Selected_From_Retrieved_But_Catalog_Missing_Count": selected_from_retrieved_but_catalog_missing,
                    "Has_Mode_Hallucination": "Y" if unique_hallucinated_ids else "N",
                    "Hallucinated_APIs": ", ".join(unique_hallucinated_ids),
                }
            )

    query_overview: List[Dict[str, Any]] = []
    mode_overview: List[Dict[str, Any]] = []
    retrieval_query_overview: List[Dict[str, Any]] = []

    grouped_mode_by_query: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    grouped_mode_by_mode: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    grouped_retrieval_by_query: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for row in mode_summary_rows:
        grouped_mode_by_query[str(row["Query_ID"])].append(row)
        grouped_mode_by_mode[str(row["Mode"])].append(row)
    for row in retrieval_summary_rows:
        grouped_retrieval_by_query[str(row["Query_ID"])].append(row)

    for query_id, rows in sorted(grouped_mode_by_query.items(), key=lambda item: _query_sort_key(item[0])):
        query_overview.append(
            {
                "Key": query_id,
                "Groups_Checked": len(rows),
                "Groups_With_Hallucinations": sum(1 for row in rows if row["Has_Mode_Hallucination"] == "Y"),
                "Hallucinated_API_Count": sum(int(row["Hallucination_Row_Count"]) for row in rows),
            }
        )

    for mode, rows in sorted(grouped_mode_by_mode.items(), key=lambda item: _mode_sort_key(item[0])):
        mode_overview.append(
            {
                "Key": mode,
                "Groups_Checked": len(rows),
                "Groups_With_Hallucinations": sum(1 for row in rows if row["Has_Mode_Hallucination"] == "Y"),
                "Hallucinated_API_Count": sum(int(row["Hallucination_Row_Count"]) for row in rows),
            }
        )

    for query_id, rows in sorted(grouped_retrieval_by_query.items(), key=lambda item: _query_sort_key(item[0])):
        retrieval_query_overview.append(
            {
                "Key": query_id,
                "Groups_Checked": len(rows),
                "Groups_With_Hallucinations": sum(1 for row in rows if row["Has_Retrieval_Hallucination"] == "Y"),
                "Hallucinated_API_Count": sum(int(row["Retrieved_Not_In_Catalog_Count"]) for row in rows),
            }
        )

    overall = {
        "root_dir": str(root_dir),
        "catalog_path": str(catalog_path),
        "run_count": len({row["Run_Dir"] for row in mode_summary_rows} | {row["Run_Dir"] for row in retrieval_summary_rows}),
        "retrieval_groups_checked": len(retrieval_summary_rows),
        "retrieval_groups_with_hallucinations": sum(1 for row in retrieval_summary_rows if row["Has_Retrieval_Hallucination"] == "Y"),
        "retrieved_api_not_in_catalog_count": sum(int(row["Retrieved_Not_In_Catalog_Count"]) for row in retrieval_summary_rows),
        "mode_groups_checked": len(mode_summary_rows),
        "mode_groups_with_hallucinations": sum(1 for row in mode_summary_rows if row["Has_Mode_Hallucination"] == "Y"),
        "mode_hallucination_row_count": sum(int(row["Hallucination_Row_Count"]) for row in mode_summary_rows),
        "selected_not_in_retrieved_count": sum(int(row["Selected_Not_In_Retrieved_Count"]) for row in mode_summary_rows),
        "selected_not_in_catalog_count": sum(int(row["Selected_Not_In_Catalog_Count"]) for row in mode_summary_rows),
        "selected_from_retrieved_but_catalog_missing_count": sum(
            int(row["Selected_From_Retrieved_But_Catalog_Missing_Count"]) for row in mode_summary_rows
        ),
        "missing_runs": missing_runs,
    }

    return {
        "overall": overall,
        "retrieval_summary_rows": retrieval_summary_rows,
        "retrieval_detail_rows": retrieval_detail_rows,
        "mode_summary_rows": mode_summary_rows,
        "mode_detail_rows": mode_detail_rows,
        "query_overview": query_overview,
        "mode_overview": mode_overview,
        "retrieval_query_overview": retrieval_query_overview,
    }


def collect_hallucination_audit(root_dir: Path, catalog_path: Path) -> Dict[str, Any]:
    return _collect_hallucination_audit_for_runs(root_dir, catalog_path, _iter_run_dirs(root_dir))


def collect_hallucination_audit_for_run(run_dir: Path, catalog_path: Path) -> Dict[str, Any]:
    return _collect_hallucination_audit_for_runs(run_dir, catalog_path, [run_dir])


def _default_output_targets(root_dir: Path) -> Tuple[Path, Path, Path]:
    if _is_query_run_dir(root_dir):
        query_id = root_dir.name.split("_", 1)[0]
        eval_dir = _find_evaluation_dir(root_dir) or (root_dir / "evaluation")
        stem = eval_dir / f"query_{query_id}_hallucination_audit"
        return stem.with_suffix(".xlsx"), stem.with_suffix(".json"), eval_dir / f"query_{query_id}_hallucination_audit_csv"
    return root_dir / "api_hallucination_audit.xlsx", root_dir / "api_hallucination_audit.json", root_dir / "api_hallucination_audit_csv"


def _write_csv(path: Path, columns: List[str], rows: List[Dict[str, Any]]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})
    return path


def _append_sheet(ws, columns: List[str], rows: List[Dict[str, Any]], highlight_key: str | None = None) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.append(columns)
    for idx in range(1, len(columns) + 1):
        ws.cell(row=1, column=idx).font = Font(bold=True)
    ws.freeze_panes = "A2"

    alert_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    for row in rows:
        ws.append([row.get(column, "") for column in columns])
        if highlight_key and row.get(highlight_key) == "Y":
            for idx in range(1, len(columns) + 1):
                ws.cell(row=ws.max_row, column=idx).fill = alert_fill

    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 60)


def write_hallucination_audit_excel(audit: Dict[str, Any], out_path: Path) -> Path:
    from openpyxl import Workbook

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    overview_ws = wb.active
    overview_ws.title = "Overview"
    overview_rows = [
        {"Metric": "Root Dir", "Value": audit["overall"]["root_dir"]},
        {"Metric": "Catalog Path", "Value": audit["overall"]["catalog_path"]},
        {"Metric": "Runs Checked", "Value": audit["overall"]["run_count"]},
        {"Metric": "Retrieval Groups Checked", "Value": audit["overall"]["retrieval_groups_checked"]},
        {"Metric": "Retrieval Groups With Hallucinations", "Value": audit["overall"]["retrieval_groups_with_hallucinations"]},
        {"Metric": "Retrieved APIs Not In Catalog", "Value": audit["overall"]["retrieved_api_not_in_catalog_count"]},
        {"Metric": "Mode Groups Checked", "Value": audit["overall"]["mode_groups_checked"]},
        {"Metric": "Mode Groups With Hallucinations", "Value": audit["overall"]["mode_groups_with_hallucinations"]},
        {"Metric": "Hallucinated Mode Rows", "Value": audit["overall"]["mode_hallucination_row_count"]},
        {"Metric": "Selected APIs Not In Retrieved", "Value": audit["overall"]["selected_not_in_retrieved_count"]},
        {"Metric": "Selected APIs Not In Catalog", "Value": audit["overall"]["selected_not_in_catalog_count"]},
        {
            "Metric": "Selected APIs From Retrieved But Catalog Missing",
            "Value": audit["overall"]["selected_from_retrieved_but_catalog_missing_count"],
        },
        {"Metric": "Missing Runs", "Value": ", ".join(audit["overall"]["missing_runs"]) if audit["overall"]["missing_runs"] else ""},
    ]
    _append_sheet(overview_ws, ["Metric", "Value"], overview_rows)

    retrieval_ws = wb.create_sheet("RetrievalAudit")
    _append_sheet(retrieval_ws, RETRIEVAL_SUMMARY_COLUMNS, audit["retrieval_summary_rows"], highlight_key="Has_Retrieval_Hallucination")

    retrieval_detail_ws = wb.create_sheet("RetrievalDetails")
    _append_sheet(retrieval_detail_ws, RETRIEVAL_DETAIL_COLUMNS, audit["retrieval_detail_rows"])

    mode_ws = wb.create_sheet("ModeAudit")
    _append_sheet(mode_ws, MODE_SUMMARY_COLUMNS, audit["mode_summary_rows"], highlight_key="Has_Mode_Hallucination")

    mode_detail_ws = wb.create_sheet("ModeDetails")
    _append_sheet(mode_detail_ws, MODE_DETAIL_COLUMNS, audit["mode_detail_rows"])

    query_ws = wb.create_sheet("ModeQueryOverview")
    _append_sheet(query_ws, OVERVIEW_COLUMNS, audit["query_overview"])

    retrieval_query_ws = wb.create_sheet("RetrievalQueryOverview")
    _append_sheet(retrieval_query_ws, OVERVIEW_COLUMNS, audit["retrieval_query_overview"])

    mode_overview_ws = wb.create_sheet("ModeOverview")
    _append_sheet(mode_overview_ws, OVERVIEW_COLUMNS, audit["mode_overview"])

    wb.save(out_path)
    return out_path


def write_hallucination_audit_csv_bundle(audit: Dict[str, Any], out_dir: Path) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    return [
        _write_csv(out_dir / "retrieval_catalog_audit.csv", RETRIEVAL_SUMMARY_COLUMNS, audit["retrieval_summary_rows"]),
        _write_csv(out_dir / "retrieval_catalog_missing_details.csv", RETRIEVAL_DETAIL_COLUMNS, audit["retrieval_detail_rows"]),
        _write_csv(out_dir / "mode_result_hallucination_audit.csv", MODE_SUMMARY_COLUMNS, audit["mode_summary_rows"]),
        _write_csv(out_dir / "mode_result_hallucination_details.csv", MODE_DETAIL_COLUMNS, audit["mode_detail_rows"]),
        _write_csv(out_dir / "mode_query_overview.csv", OVERVIEW_COLUMNS, audit["query_overview"]),
        _write_csv(out_dir / "retrieval_query_overview.csv", OVERVIEW_COLUMNS, audit["retrieval_query_overview"]),
        _write_csv(out_dir / "mode_overview.csv", OVERVIEW_COLUMNS, audit["mode_overview"]),
        _write_csv(
            out_dir / "overall_summary.csv",
            ["Metric", "Value"],
            [
                {"Metric": "Root Dir", "Value": audit["overall"]["root_dir"]},
                {"Metric": "Catalog Path", "Value": audit["overall"]["catalog_path"]},
                {"Metric": "Runs Checked", "Value": audit["overall"]["run_count"]},
                {"Metric": "Retrieval Groups Checked", "Value": audit["overall"]["retrieval_groups_checked"]},
                {"Metric": "Retrieval Groups With Hallucinations", "Value": audit["overall"]["retrieval_groups_with_hallucinations"]},
                {"Metric": "Retrieved APIs Not In Catalog", "Value": audit["overall"]["retrieved_api_not_in_catalog_count"]},
                {"Metric": "Mode Groups Checked", "Value": audit["overall"]["mode_groups_checked"]},
                {"Metric": "Mode Groups With Hallucinations", "Value": audit["overall"]["mode_groups_with_hallucinations"]},
                {"Metric": "Hallucinated Mode Rows", "Value": audit["overall"]["mode_hallucination_row_count"]},
                {"Metric": "Selected APIs Not In Retrieved", "Value": audit["overall"]["selected_not_in_retrieved_count"]},
                {"Metric": "Selected APIs Not In Catalog", "Value": audit["overall"]["selected_not_in_catalog_count"]},
                {
                    "Metric": "Selected APIs From Retrieved But Catalog Missing",
                    "Value": audit["overall"]["selected_from_retrieved_but_catalog_missing_count"],
                },
                {"Metric": "Missing Runs", "Value": ", ".join(audit["overall"]["missing_runs"]) if audit["overall"]["missing_runs"] else ""},
            ],
        ),
    ]


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit retrieval and mode-output API hallucinations against the catalog and retrieved candidates.")
    parser.add_argument(
        "--root-dir",
        type=Path,
        default=Path("results/logs/RUNS_APR_20"),
        help="Root directory containing run folders or provider/model subdirectories.",
    )
    parser.add_argument(
        "--catalog-path",
        type=Path,
        default=catalog_path(with_qos=False),
        help="Catalog JSONL used as the source of truth for valid API ids.",
    )
    parser.add_argument(
        "--output-json",
        type=Path,
        default=None,
        help="Output JSON path. Defaults to <root-dir>/api_hallucination_audit.json",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=None,
        help="Output workbook path. Defaults to <root-dir>/api_hallucination_audit.xlsx",
    )
    parser.add_argument(
        "--output-csv-dir",
        type=Path,
        default=None,
        help="Output CSV directory. Defaults to <root-dir>/api_hallucination_audit_csv when XLSX is unavailable.",
    )
    args = parser.parse_args()

    root_dir = args.root_dir.expanduser().resolve()
    catalog_path = args.catalog_path.expanduser().resolve()
    if not root_dir.exists():
        raise FileNotFoundError(f"Root directory not found: {root_dir}")
    if not catalog_path.exists():
        raise FileNotFoundError(f"Catalog path not found: {catalog_path}")

    default_xlsx, default_json, default_csv_dir = _default_output_targets(root_dir)
    output_json = args.output_json.expanduser().resolve() if args.output_json else default_json
    output_xlsx = args.output_xlsx.expanduser().resolve() if args.output_xlsx else default_xlsx
    output_csv_dir = args.output_csv_dir.expanduser().resolve() if args.output_csv_dir else default_csv_dir

    audit = collect_hallucination_audit(root_dir, catalog_path)
    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(audit, indent=2, ensure_ascii=False), encoding="utf-8")

    wrote_excel = False
    try:
        write_hallucination_audit_excel(audit, output_xlsx)
        wrote_excel = True
    except ModuleNotFoundError:
        write_hallucination_audit_csv_bundle(audit, output_csv_dir)

    overall = audit["overall"]
    print(f"Checked {overall['run_count']} runs under {root_dir}")
    print(f"Retrieval groups checked: {overall['retrieval_groups_checked']}")
    print(f"Retrieval groups with hallucinations: {overall['retrieval_groups_with_hallucinations']}")
    print(f"Retrieved APIs not in catalog: {overall['retrieved_api_not_in_catalog_count']}")
    print(f"Mode groups checked: {overall['mode_groups_checked']}")
    print(f"Mode groups with hallucinations: {overall['mode_groups_with_hallucinations']}")
    print(f"Hallucinated mode rows: {overall['mode_hallucination_row_count']}")
    print(f"Selected APIs not in retrieved set: {overall['selected_not_in_retrieved_count']}")
    print(f"Selected APIs not in catalog: {overall['selected_not_in_catalog_count']}")
    print(f"Selected APIs from retrieved set but missing in catalog: {overall['selected_from_retrieved_but_catalog_missing_count']}")
    if wrote_excel:
        print(f"Wrote Excel audit to {output_xlsx}")
    else:
        print("openpyxl not installed; wrote CSV audit bundle instead")
        print(f"Wrote CSV audit bundle to {output_csv_dir}")
    print(f"Wrote JSON audit to {output_json}")


if __name__ == "__main__":
    main()
