from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

MODE_ORDER = ["no_qos", "qos_pure_llm", "qos_topsis", "qos_hybrid"]
MODE_INDEX = {name: idx for idx, name in enumerate(MODE_ORDER)}

SUMMARY_COLUMNS = [
    "Run_Dir",
    "Query_ID",
    "Sub_Task",
    "Mode",
    "Subtask_Purpose",
    "Candidate_Count",
    "Unique_API_Count",
    "Duplicate_API_Count",
    "Duplicate_Row_Count",
    "Max_Repeat_Count",
    "Has_Duplicates",
    "Duplicated_APIs",
]

DUPLICATE_COLUMNS = [
    "Run_Dir",
    "Query_ID",
    "Sub_Task",
    "Mode",
    "Subtask_Purpose",
    "Selected_API",
    "Repeat_Count",
    "Mode_Ranks",
    "Retrieved_Ranks",
    "Functional_Match_Values",
    "Comments",
]

OVERVIEW_COLUMNS = [
    "Key",
    "Groups_Checked",
    "Groups_With_Duplicates",
    "Duplicate_API_Count",
    "Duplicate_Row_Count",
]


def _query_sort_key(value: str) -> Tuple[int, str]:
    match = re.search(r"q(\d+)", str(value), flags=re.IGNORECASE)
    if match:
        return int(match.group(1)), str(value)
    return 9999, str(value)


def _subtask_sort_key(value: Any) -> Tuple[int, str]:
    text = str(value)
    return (int(text), text) if text.isdigit() else (9999, text)


def _mode_sort_key(value: str) -> Tuple[int, str]:
    text = str(value)
    return MODE_INDEX.get(text, 9999), text


def _int_sort(values: Iterable[Any]) -> List[str]:
    def _key(value: Any) -> Tuple[int, str]:
        text = str(value)
        return (int(text), text) if text.isdigit() else (9999, text)

    return [str(v) for v in sorted(values, key=_key)]


def _safe_load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _is_query_run_dir(path: Path) -> bool:
    return path.is_dir() and (path / "0_decomposer.json").exists()


def _iter_run_dirs(root_dir: Path) -> Iterable[Path]:
    if _is_query_run_dir(root_dir):
        yield root_dir
        return
    for path in sorted(root_dir.iterdir(), key=lambda p: _query_sort_key(p.name)):
        if path.is_dir():
            yield path


def _find_evaluation_dir(run_dir: Path) -> Path | None:
    for candidate in ("evaluation", "functional_match_eval"):
        eval_dir = run_dir / candidate
        if eval_dir.exists():
            return eval_dir
    return None


def _find_rows_json(run_dir: Path) -> Path | None:
    eval_dir = _find_evaluation_dir(run_dir)
    if eval_dir is None:
        return None
    matches = sorted(eval_dir.glob("query_*_candidate_api_rankings_rows.json"))
    return matches[0] if matches else None


def _collect_group_summary(
    *,
    run_dir_name: str,
    query_id: str,
    subtask_id: str,
    mode: str,
    rows: List[Dict[str, Any]],
) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    cleaned_api_ids = [str(row.get("Selected_API", "")).strip() for row in rows if str(row.get("Selected_API", "")).strip()]
    counts = Counter(cleaned_api_ids)
    duplicate_counts = {api_id: count for api_id, count in counts.items() if count > 1}
    duplicate_api_count = len(duplicate_counts)
    duplicate_row_count = sum(count - 1 for count in duplicate_counts.values())
    max_repeat_count = max(counts.values(), default=0)
    subtask_purpose = str(rows[0].get("Subtask_Purpose", "")) if rows else ""

    summary = {
        "Run_Dir": run_dir_name,
        "Query_ID": query_id,
        "Sub_Task": subtask_id,
        "Mode": mode,
        "Subtask_Purpose": subtask_purpose,
        "Candidate_Count": len(rows),
        "Unique_API_Count": len(counts),
        "Duplicate_API_Count": duplicate_api_count,
        "Duplicate_Row_Count": duplicate_row_count,
        "Max_Repeat_Count": max_repeat_count,
        "Has_Duplicates": "Y" if duplicate_counts else "N",
        "Duplicated_APIs": ", ".join(f"{api_id} x{count}" for api_id, count in sorted(duplicate_counts.items())),
    }

    duplicate_details: List[Dict[str, Any]] = []
    for api_id, repeat_count in sorted(duplicate_counts.items()):
        matching_rows = [row for row in rows if str(row.get("Selected_API", "")).strip() == api_id]
        mode_ranks = _int_sort(row.get("Mode Rank", "") for row in matching_rows)
        retrieved_ranks = _int_sort(row.get("Retrieved Rank", "") for row in matching_rows)
        functional_match_values = [str(row.get("Functional Match (0/1)", "")) for row in matching_rows]
        unique_comments = []
        seen_comments = set()
        for row in matching_rows:
            comment = str(row.get("Comments", "")).strip()
            if comment and comment not in seen_comments:
                seen_comments.add(comment)
                unique_comments.append(comment)

        duplicate_details.append(
            {
                "Run_Dir": run_dir_name,
                "Query_ID": query_id,
                "Sub_Task": subtask_id,
                "Mode": mode,
                "Subtask_Purpose": subtask_purpose,
                "Selected_API": api_id,
                "Repeat_Count": repeat_count,
                "Mode_Ranks": ", ".join(mode_ranks),
                "Retrieved_Ranks": ", ".join(retrieved_ranks),
                "Functional_Match_Values": ", ".join(functional_match_values),
                "Comments": " | ".join(unique_comments[:3]),
            }
        )

    return summary, duplicate_details


def _collect_duplicate_audit_for_runs(root_dir: Path, run_dirs: Iterable[Path]) -> Dict[str, Any]:
    summary_rows: List[Dict[str, Any]] = []
    duplicate_rows: List[Dict[str, Any]] = []
    missing_runs: List[str] = []

    for run_dir in run_dirs:
        rows_path = _find_rows_json(run_dir)
        if rows_path is None:
            missing_runs.append(run_dir.name)
            continue

        rows = _safe_load_json(rows_path)
        if not isinstance(rows, list):
            continue

        grouped: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
        for row in rows:
            query_id = str(row.get("Query_ID", "")).strip()
            subtask_id = str(row.get("Sub Task", "")).strip()
            mode = str(row.get("Mode", "")).strip()
            grouped[(query_id, subtask_id, mode)].append(row)

        for (query_id, subtask_id, mode), group_rows in sorted(
            grouped.items(),
            key=lambda item: (_query_sort_key(item[0][0]), _subtask_sort_key(item[0][1]), _mode_sort_key(item[0][2])),
        ):
            summary, details = _collect_group_summary(
                run_dir_name=run_dir.name,
                query_id=query_id,
                subtask_id=subtask_id,
                mode=mode,
                rows=group_rows,
            )
            summary_rows.append(summary)
            duplicate_rows.extend(details)

    query_overview: List[Dict[str, Any]] = []
    mode_overview: List[Dict[str, Any]] = []

    grouped_by_query: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    grouped_by_mode: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in summary_rows:
        grouped_by_query[str(row["Query_ID"])].append(row)
        grouped_by_mode[str(row["Mode"])].append(row)

    for query_id, rows in sorted(grouped_by_query.items(), key=lambda item: _query_sort_key(item[0])):
        query_overview.append(
            {
                "Key": query_id,
                "Groups_Checked": len(rows),
                "Groups_With_Duplicates": sum(1 for row in rows if row["Has_Duplicates"] == "Y"),
                "Duplicate_API_Count": sum(int(row["Duplicate_API_Count"]) for row in rows),
                "Duplicate_Row_Count": sum(int(row["Duplicate_Row_Count"]) for row in rows),
            }
        )

    for mode, rows in sorted(grouped_by_mode.items(), key=lambda item: _mode_sort_key(item[0])):
        mode_overview.append(
            {
                "Key": mode,
                "Groups_Checked": len(rows),
                "Groups_With_Duplicates": sum(1 for row in rows if row["Has_Duplicates"] == "Y"),
                "Duplicate_API_Count": sum(int(row["Duplicate_API_Count"]) for row in rows),
                "Duplicate_Row_Count": sum(int(row["Duplicate_Row_Count"]) for row in rows),
            }
        )

    overall = {
        "root_dir": str(root_dir),
        "run_count": len({row["Run_Dir"] for row in summary_rows}),
        "group_count": len(summary_rows),
        "groups_with_duplicates": sum(1 for row in summary_rows if row["Has_Duplicates"] == "Y"),
        "duplicate_api_count": sum(int(row["Duplicate_API_Count"]) for row in summary_rows),
        "duplicate_row_count": sum(int(row["Duplicate_Row_Count"]) for row in summary_rows),
        "missing_runs": missing_runs,
    }

    return {
        "overall": overall,
        "summary_rows": summary_rows,
        "duplicate_rows": duplicate_rows,
        "query_overview": query_overview,
        "mode_overview": mode_overview,
    }


def collect_duplicate_audit(root_dir: Path) -> Dict[str, Any]:
    return _collect_duplicate_audit_for_runs(root_dir, _iter_run_dirs(root_dir))


def collect_duplicate_audit_for_run(run_dir: Path) -> Dict[str, Any]:
    return _collect_duplicate_audit_for_runs(run_dir, [run_dir])


def _default_output_targets(root_dir: Path) -> Tuple[Path, Path, Path]:
    if _is_query_run_dir(root_dir):
        query_id = root_dir.name.split("_", 1)[0]
        eval_dir = _find_evaluation_dir(root_dir) or (root_dir / "evaluation")
        stem = eval_dir / f"query_{query_id}_duplicate_audit"
        return stem.with_suffix(".xlsx"), stem.with_suffix(".json"), eval_dir / f"query_{query_id}_duplicate_audit_csv"
    return root_dir / "api_duplicate_audit.xlsx", root_dir / "api_duplicate_audit.json", root_dir / "api_duplicate_audit_csv"


def _append_sheet(ws, columns: List[str], rows: List[Dict[str, Any]], highlight_duplicates: bool = False) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.append(columns)
    for idx in range(1, len(columns) + 1):
        ws.cell(row=1, column=idx).font = Font(bold=True)
    ws.freeze_panes = "A2"

    alert_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    for row in rows:
        ws.append([row.get(column, "") for column in columns])
        if highlight_duplicates and row.get("Has_Duplicates") == "Y":
            for idx in range(1, len(columns) + 1):
                ws.cell(row=ws.max_row, column=idx).fill = alert_fill

    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 60)


def write_duplicate_audit_excel(audit: Dict[str, Any], out_path: Path) -> Path:
    from openpyxl import Workbook

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    overview_ws = wb.active
    overview_ws.title = "Overview"
    overview_rows = [
        {"Metric": "Root Dir", "Value": audit["overall"]["root_dir"]},
        {"Metric": "Runs Checked", "Value": audit["overall"]["run_count"]},
        {"Metric": "Groups Checked", "Value": audit["overall"]["group_count"]},
        {"Metric": "Groups With Duplicates", "Value": audit["overall"]["groups_with_duplicates"]},
        {"Metric": "Duplicate API Count", "Value": audit["overall"]["duplicate_api_count"]},
        {"Metric": "Duplicate Row Count", "Value": audit["overall"]["duplicate_row_count"]},
        {"Metric": "Missing Runs", "Value": ", ".join(audit["overall"]["missing_runs"]) if audit["overall"]["missing_runs"] else ""},
    ]
    _append_sheet(overview_ws, ["Metric", "Value"], overview_rows)

    summary_ws = wb.create_sheet("SubtaskModeAudit")
    _append_sheet(summary_ws, SUMMARY_COLUMNS, audit["summary_rows"], highlight_duplicates=True)

    duplicates_ws = wb.create_sheet("DuplicateDetails")
    _append_sheet(duplicates_ws, DUPLICATE_COLUMNS, audit["duplicate_rows"])

    query_ws = wb.create_sheet("QueryOverview")
    _append_sheet(query_ws, OVERVIEW_COLUMNS, audit["query_overview"])

    mode_ws = wb.create_sheet("ModeOverview")
    _append_sheet(mode_ws, OVERVIEW_COLUMNS, audit["mode_overview"])

    wb.save(out_path)
    return out_path


def _write_csv(path: Path, columns: List[str], rows: List[Dict[str, Any]]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})
    return path


def write_duplicate_audit_csv_bundle(audit: Dict[str, Any], out_dir: Path) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    written = [
        _write_csv(out_dir / "subtask_mode_audit.csv", SUMMARY_COLUMNS, audit["summary_rows"]),
        _write_csv(out_dir / "duplicate_details.csv", DUPLICATE_COLUMNS, audit["duplicate_rows"]),
        _write_csv(out_dir / "query_overview.csv", OVERVIEW_COLUMNS, audit["query_overview"]),
        _write_csv(out_dir / "mode_overview.csv", OVERVIEW_COLUMNS, audit["mode_overview"]),
        _write_csv(
            out_dir / "overall_summary.csv",
            ["Metric", "Value"],
            [
                {"Metric": "Root Dir", "Value": audit["overall"]["root_dir"]},
                {"Metric": "Runs Checked", "Value": audit["overall"]["run_count"]},
                {"Metric": "Groups Checked", "Value": audit["overall"]["group_count"]},
                {"Metric": "Groups With Duplicates", "Value": audit["overall"]["groups_with_duplicates"]},
                {"Metric": "Duplicate API Count", "Value": audit["overall"]["duplicate_api_count"]},
                {"Metric": "Duplicate Row Count", "Value": audit["overall"]["duplicate_row_count"]},
                {"Metric": "Missing Runs", "Value": ", ".join(audit["overall"]["missing_runs"]) if audit["overall"]["missing_runs"] else ""},
            ],
        ),
    ]
    return written


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit duplicate Selected_API values within each query/subtask/mode group.")
    parser.add_argument(
        "--root-dir",
        type=Path,
        default=Path("results/logs/RUNS_APR_20/mistral_mistral-small-latest"),
        help="Directory containing query run folders.",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=None,
        help="Output workbook path. Defaults to <root-dir>/api_duplicate_audit.xlsx",
    )
    parser.add_argument(
        "--output-json",
        type=Path,
        default=None,
        help="Output JSON path. Defaults to <root-dir>/api_duplicate_audit.json",
    )
    parser.add_argument(
        "--output-csv-dir",
        type=Path,
        default=None,
        help="Optional CSV output directory. Defaults to <root-dir>/api_duplicate_audit_csv when XLSX is unavailable.",
    )
    args = parser.parse_args()

    root_dir = args.root_dir.expanduser().resolve()
    if not root_dir.exists():
        raise FileNotFoundError(f"Root directory not found: {root_dir}")

    default_xlsx, default_json, default_csv_dir = _default_output_targets(root_dir)
    output_xlsx = args.output_xlsx.expanduser().resolve() if args.output_xlsx else default_xlsx
    output_json = args.output_json.expanduser().resolve() if args.output_json else default_json
    output_csv_dir = args.output_csv_dir.expanduser().resolve() if args.output_csv_dir else default_csv_dir

    audit = collect_duplicate_audit(root_dir)
    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(audit, indent=2, ensure_ascii=False), encoding="utf-8")

    wrote_excel = False
    try:
        write_duplicate_audit_excel(audit, output_xlsx)
        wrote_excel = True
    except ModuleNotFoundError:
        write_duplicate_audit_csv_bundle(audit, output_csv_dir)

    overall = audit["overall"]
    print(f"Checked {overall['run_count']} runs under {root_dir}")
    print(f"Checked {overall['group_count']} query/subtask/mode groups")
    print(f"Groups with duplicates: {overall['groups_with_duplicates']}")
    print(f"Duplicate API ids counted: {overall['duplicate_api_count']}")
    print(f"Duplicate extra rows counted: {overall['duplicate_row_count']}")
    if wrote_excel:
        print(f"Wrote Excel audit to {output_xlsx}")
    else:
        print("openpyxl not installed; wrote CSV audit bundle instead")
        print(f"Wrote CSV audit bundle to {output_csv_dir}")
    print(f"Wrote JSON audit to {output_json}")


if __name__ == "__main__":
    main()
